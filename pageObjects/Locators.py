import os
import time
import datetime

from selenium import webdriver
from selenium.common import NoSuchElementException, TimeoutException, StaleElementReferenceException
from selenium.webdriver import ActionChains, Keys

from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.remote import webelement
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.wait import WebDriverWait
from utilities import XLUtils

from utilities.customLogger import LogGen
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl
import os
import shutil


class LoginPage:
    textbox_username_xpath = "//input[@placeholder='Email or username']"
    textbox_password_xpath = "//input[@placeholder='Password']"
    button_login_xpath = "//input[@type='submit']"
    checkbox_rememberme_xpath = "//div[@class='field remember-me']//div[@class='remember-me uwa-checkbox-content uwa-toggle-content uwa-input-content uwa-icon']"
    link_iamnot_xpath = "//a[contains(.,'I am not')]"
    button_profile_xpath = "//div[@class='profile-picture']"
    button_logout_xpath = "//span[contains(text(),'Log out')]"
    switchUser_reset_xpath = "//a[contains(.,'I am not')]"
    maximize_xpath = "//div[contains(@class, 'moduleWrapper clearfix ifwe-tabview')]//div[@class='wp-tabview-panel']//div[@class='widget-dd-menu dropdown-menu dropdown-menu-root dropdown dropdown-root']//child::span[@class='maximize-icon fonticon fonticon-resize-full']"
    restore_xpath = "//div[contains(@class, 'moduleWrapper clearfix ifwe-tabview')]//div[@class='wp-tabview-panel']//div[@class='widget-dd-menu dropdown-menu dropdown-menu-root dropdown dropdown-root']//child::span[@class='maximize-icon fonticon fonticon-resize-small']"

    # constructer to initialise driver
    def __init__(self, driver):
        self.driver = driver

    # # logger = LogGen.loggen()

    # Action Methods of LoginPage
    def setUserName(self, username):
        try:
            try:
                WebDriverWait(self.driver, 5).until(expected_conditions.element_to_be_clickable(
                    (By.XPATH, self.link_iamnot_xpath))).click()

                # self.driver.find_element(By.XPATH,'link_iamnot_xpath')
                #
                # ActionChains(self.driver).move_to_element(self.link_iamnot_xpath).click(self.link_iamnot_xpath).perform()

                WebDriverWait(self.driver, 10).until(expected_conditions.presence_of_element_located(
                    (By.XPATH, self.textbox_username_xpath))).clear()
                WebDriverWait(self.driver, 10).until(expected_conditions.presence_of_element_located(
                    (By.XPATH, self.textbox_username_xpath))).send_keys(username)
            except:
                WebDriverWait(self.driver, 10).until(expected_conditions.presence_of_element_located(
                    (By.XPATH, self.textbox_username_xpath))).clear()
                WebDriverWait(self.driver, 10).until(expected_conditions.presence_of_element_located(
                    (By.XPATH, self.textbox_username_xpath))).send_keys(username)

        except Exception as e:
            raise e

    def setPassword(self, password):
        WebDriverWait(self.driver, 10).until(expected_conditions.element_to_be_clickable(
            (By.XPATH, self.textbox_password_xpath))).clear()
        WebDriverWait(self.driver, 10).until(expected_conditions.element_to_be_clickable(
            (By.XPATH, self.textbox_password_xpath))).send_keys(password)

    def clickLogin(self):
        WebDriverWait(self.driver, 10).until(expected_conditions.element_to_be_clickable(
            (By.XPATH, self.button_login_xpath))).click()

    def uncheckRememberMe(self):
        self.driver.find_element(By.XPATH, self.checkbox_rememberme_xpath).click()

    def clickIAmNot(self):
        WebDriverWait(self.driver, 10).until(expected_conditions.element_to_be_clickable(
            (By.XPATH, self.link_iamnot_xpath))).click()

    def clickProfile(self):
        WebDriverWait(self.driver, 10).until(expected_conditions.element_to_be_clickable(
            (By.XPATH, self.button_profile_xpath))).click()

    def clickLogout(self):
        WebDriverWait(self.driver, 10).until(expected_conditions.element_to_be_clickable(
            (By.XPATH, self.button_logout_xpath))).click()

    def switchUser(self):
        WebDriverWait(self.driver, 10).until(expected_conditions.element_to_be_clickable(
            (By.XPATH, self.switchUser_reset_xpath))).click()


class DashboardAndTabs:  # all dashboard tabs locators
    link_projectinsights_tab_xpath = "//span[contains(@class, 'title') and text() = 'Project Insights']"
    link_documentmanagement_tab_xpath = "//span[contains(@class, 'title') and text() = 'Document Management']"
    link_workflowmanagement_tab_xpath = "//span[contains(@class, 'title') and text() = 'Workflow Management']"
    link_mailmanagement_tab_xpath = "//span[contains(@class, 'title') and text() = 'Mail Management']"
    link_alltasksview_tab_xpath = "//span[contains(@class, 'title') and text() = 'All Tasks View']"
    title_documentreg_xpath = "//div[contains(@class, 'moduleHeader__title') and text() = 'Document Register']"
    # menu_dropdown_xpath_1 = "//div[contains(@class, 'moduleWrapper clearfix ifwe-tabview')]//div[@class='wp-tabview-panel']//div[@id='m_9sMywGd0g06qZr2-U046']//child::span[@class='widget-menu-icon ifwe-action-icon fonticon fonticon-down-open clickable']"
    # menu_dropdown_xpath = "//div[contains(@class, 'moduleWrapper clearfix ifwe-tabview')]//div[@class='wp-tabview-panel']//child::div[contains(@class,'moduleHeader__title') and text()='Document Register']//following-sibling::div//child::span[@class='widget-menu-icon ifwe-action-icon fonticon fonticon-down-open clickable']"
    maximize_xpath = "//div[contains(@class, 'moduleWrapper clearfix ifwe-tabview')]//div[@class='wp-tabview-panel']//div[@class='widget-dd-menu dropdown-menu dropdown-menu-root dropdown dropdown-root']//child::span[@class='maximize-icon fonticon fonticon-resize-full']"
    restore_xpath = "//div[contains(@class, 'moduleWrapper clearfix ifwe-tabview')]//div[@class='wp-tabview-panel']//div[@class='widget-dd-menu dropdown-menu dropdown-menu-root dropdown dropdown-root']//child::span[@class='maximize-icon fonticon fonticon-resize-small']"

    # constructer to initialise driver
    def __init__(self, driver):
        self.driver = driver

    # # logger = LogGen.loggen()

    # Action Methods of DashboardTabs
    def dashboardselection(self, dashboard):
        try:
            WebDriverWait(self.driver, 300).until(expected_conditions.presence_of_element_located(
                (By.XPATH, "//span[@class='topbar-app-name']")))
            WebDriverWait(self.driver, 300).until(expected_conditions.presence_of_element_located(
                (By.XPATH, "//span[contains(@class,'topbar-app-name') and text()='" + dashboard + "']")))
            # # self.logger.info("** Dashboard is selected **")
        except NoSuchElementException:
            try:
                # click on Dashboards and cockpit list menu
                WebDriverWait(self.driver, 30).until(expected_conditions.element_to_be_clickable(
                    (
                        By.XPATH,
                        "//span[@class='wp-panel-button fonticon fonticon-menu new-dashboard-menu-open-btn inactive']"))).click()
                # # self.logger.info("** clicked on Dashboard and cockpit list menu **")
                try:
                    WebDriverWait(self.driver, 30).until(expected_conditions.element_to_be_clickable(
                        (By.XPATH,
                         "//div[@class='dashboard-menu-list-item-text']/p[text()='" + dashboard + "']"))).click()
                    # # self.logger.info("** clicked on Dashboard **")

                except Exception as e:
                    # # self.logger.info("** Fail to click on Dashboard **")
                    raise e
            except Exception as e:
                # # self.logger.info("** Fail **")
                raise e

            WebDriverWait(self.driver, 10).until(expected_conditions.element_to_be_clickable(
                (By.XPATH,
                 "//span[@class='fonticon fonticon-close new-dashboard-menu-close-btn']"))).click()
            WebDriverWait(self.driver, 10).until(expected_conditions.invisibility_of_element_located(
                (By.XPATH, "//span[@class='fonticon fonticon-close new-dashboard-menu-close-btn']")))

    def clickprojectinsightstab(self):
        WebDriverWait(self.driver, 10).until(expected_conditions.element_to_be_clickable(
            (By.XPATH, self.link_projectinsights_tab_xpath))).click()

    def clickdocumentmgtab(self):
        WebDriverWait(self.driver, 10).until(expected_conditions.element_to_be_clickable(
            (By.XPATH, self.link_documentmanagement_tab_xpath))).click()

    def clickworkflowmgtab(self):
        WebDriverWait(self.driver, 10).until(expected_conditions.element_to_be_clickable(
            (By.XPATH, self.link_workflowmanagement_tab_xpath))).click()

    def clickmailmgtab(self):
        WebDriverWait(self.driver, 10).until(expected_conditions.element_to_be_clickable(
            (By.XPATH, self.link_mailmanagement_tab_xpath))).click()

    def clickalltasksviewtab(self):
        WebDriverWait(self.driver, 10).until(expected_conditions.element_to_be_clickable(
            (By.XPATH, self.link_alltasksview_tab_xpath))).click()

    def click_menu_dropdown(self, widget_name):
        WebDriverWait(self.driver, 60).until(expected_conditions.element_to_be_clickable(
            (By.XPATH,
             "//div[contains(@class, 'moduleWrapper clearfix ifwe-tabview')]//div[@class='wp-tabview-panel']//child::div[contains(@class,'moduleHeader__title') and text()='" + widget_name + "']//following-sibling::div//child::span[@class='widget-menu-icon ifwe-action-icon fonticon fonticon-down-open clickable']"))).click()
        time.sleep(3)

    def close_menu_dropdown(self, widget_name):
        WebDriverWait(self.driver, 60).until(expected_conditions.element_to_be_clickable(
            (By.XPATH,
             "//div[contains(@class, 'moduleWrapper clearfix ifwe-tabview')]//div[@class='wp-tabview-panel']//child::div[contains(@class,'moduleHeader__title') and text()='" + widget_name + "']//following-sibling::div//child::span[@class='widget-menu-icon ifwe-action-icon fonticon fonticon-down-open clickable active']"))).click()

    def maximize_widget(self, widget_name):
        '''
        Method to maximise the widget
        '''
        try:
            maximize = (WebDriverWait(self.driver, 10).until(expected_conditions.element_to_be_clickable(
                (By.XPATH, self.maximize_xpath))))
            maximize.click()
            time.sleep(2)
            # # self.logger.info("** widget is maximized **")
        except:
            WebDriverWait(self.driver, 30).until(expected_conditions.element_to_be_clickable(
                (By.XPATH, self.restore_xpath)))
            # # self.logger.info("** widget is already maximized **")
            WebDriverWait(self.driver, 30).until(expected_conditions.element_to_be_clickable(
                (By.XPATH,
                 "//div[contains(@class, 'moduleWrapper clearfix ifwe-tabview')]//div[@class='wp-tabview-panel']//child::div[contains(@class,'moduleHeader__title') and text()='" + widget_name + "']//following-sibling::div//child::span[@class='widget-menu-icon ifwe-action-icon fonticon fonticon-down-open clickable active']"))).click()

    def restore_widget(self):
        '''
        Method to restore the widget
        '''
        try:
            WebDriverWait(self.driver, 10).until(expected_conditions.element_to_be_clickable(
                (By.XPATH, self.restore_xpath))).click()
        except:
            pass


class documentRegisterAndWorkflow:  # all document register and workflow creation locators

    iframeNMDocumentCntrl_xpath = "//div[@id='7zAw5w30gCwH0QgEjm01']//iframe[@id='frame-9sMywGd0g06qZr2-U046']"
    # iframeNMDocumentCntrl_xpath = "//iframe[contains(@id,'frame-9sMywGd0g06qZr2-U055')]"
    tabDocumentUpload_xpath = "//a[contains(.,'Document Upload')]"
    tabTemporaryDocument_xpath = "//a[contains(.,'Temporary Document')]"
    tabRegisteredDocument_xpath = "//a[contains(.,'Registered Document')]"
    tabLegacyDocument_xpath = "//a[contains(.,'Legacy Document')]"
    # select_files_xpath = "//div[@id='mass_upload']//div[contains(.,'Select File/s *')]//child::input[@id='input-19']"
    select_files_xpath = "//div[@id='mass_upload']//label[contains(text(), 'Select File')]"
    clear_file_path = "//button[@class='v-icon notranslate v-icon--link mdi mdi-close theme--light']"
    complete_upload = "//div[@role='progressbar']//i[contains(.,'100%')]"
    bt_inline_edit_du = "//button[@class='v-icon notranslate v-icon--link mdi mdi-pencil theme--light']"
    bt_mass_edit_du = "//div[@class='v-speed-dial v-speed-dial--bottom v-speed-dial--left v-speed-dial--direction-top']"
    bt_mass_edit_1 = "//button[@class='v-btn v-btn--is-elevated v-btn--fab v-btn--has-bg v-btn--round theme--dark v-size--small green']"
    mass_edit_form = "//div[@class='v-card__title']//span[contains(@class, 'headline') and text() = 'Mass Edit']"
    text_ms_title_field = "//label[text() = 'Title']/../input[@type='text']"
    drp_ms_wo_field = "//label[text() = 'Work Order/Service Order']/.."
    ms_wo_drp = "// div[@class='v-menu__content theme--light v-menu__content--fixed menuable__content__active']//div[@class='v-list v-select-list v-sheet theme--light theme--light']"
    drp_ms_status_field = "//label[text() = 'Status']/.."
    ms_status_drp = "// div[@class='v-menu__content theme--light v-menu__content--fixed menuable__content__active']"
    drp_ms_stage_field = "//label[text() = 'Stage']/.."
    ms_stage_drp = "// div[@class='v-menu__content theme--light v-menu__content--fixed menuable__content__active']"
    bt_update_all = "//span[contains(@class, 'v-btn__content') and text() = 'UPDATE ALL']"
    bt_test = "//button[@class='mb-2 v-btn v-btn--is-elevated v-btn--has-bg theme--dark v-size--default primary']/span[text()='TEST']"
    bt_save = "//button[@class='mb-2 v-btn v-btn--is-elevated v-btn--has-bg theme--light v-size--default success']/span[text()='SAVE']"
    bt_plus_menu = "//i[@class='v-icon notranslate mdi mdi-plus theme--dark']"
    bt_create_workflow = "//div[@style='transition-delay: 0.2s;']//div[@class='v-responsive__content']"
    wf_template_search = "//div[(@class='fa fa-search search') and (@linked='routeTemplateOID')]"
    route_tmp_page = "//div[@class=' search-content-set-detail content-set-detail set-detail']"
    bt_route_ok = "//button[@id='id_in_app_ok' and text() = 'OK']"
    wf_title = "//input[@name='IMP_WFTitle']"
    bt_wf_create = "//div[starts-with(@id,'formdivision')]//following-sibling::div/button[@class='btn-primary btn hd-form-submit-hover hd-form-submit-focus']"
    pg_route_mg = "//div[contains(@class,'moduleHeader__title') and text() = 'ENOVIA - Route Management - My Routes (1)']"
    iframerouteMg_xpath = "//div[@class='moduleWrapper']//iframe[starts-with(@id, 'frame-preview-')]"
    created_wf = "//div[@class='wux-layouts-datagridview-tweaker-container']/div[contains(text(), 'WF-')]"
    lnk_start_wf = "//div[@id='channel1']/div[@class='maturity-state-container']//a[contains(text(), 'Start')]"
    awaiting_approval_xpath = "//div[@id='channel1']/div[@class='maturity-state-container']//span[@class='rt-activity-state' and @title='Awaiting Approval']"

    # constructer to initialise diver
    def __init__(self, driver):
        self.driver = driver

    # # logger = LogGen.loggen()

    def navigate_to_tab_document_upload(self):
        '''
                Method to navigate to the Document tab in Document Register widget
                '''
        try:
            self.driver.switch_to.default_content()
            frame1 = WebDriverWait(self.driver, 10).until(expected_conditions.presence_of_element_located(
                (By.XPATH, self.iframeNMDocumentCntrl_xpath)))
            self.driver.switch_to.frame(frame1)
            WebDriverWait(self.driver, 10).until(expected_conditions.element_to_be_clickable(
                (By.XPATH, self.tabDocumentUpload_xpath))).click()

        except Exception as e:
            raise e

    def navigate_to_tab_registered_document(self):
        '''
                Method to navigate to the Registered Document tab in Document Register widget
                '''
        try:
            self.driver.switch_to.default_content()
            frame1 = WebDriverWait(self.driver, 500).until(expected_conditions.presence_of_element_located(
                (By.XPATH, self.iframeNMDocumentCntrl_xpath)))
            self.driver.switch_to.frame(frame1)
            WebDriverWait(self.driver, 40).until(expected_conditions.element_to_be_clickable(
                (By.XPATH, self.tabRegisteredDocument_xpath))).click()

        except Exception as e:
            raise e

    def select_file(self, folder_path):
        '''
                Method to select files from local
                '''
        try:
            self.driver.switch_to.default_content()
            frame1 = WebDriverWait(self.driver, 10).until(expected_conditions.presence_of_element_located(
                (By.XPATH, self.iframeNMDocumentCntrl_xpath)))
            self.driver.switch_to.frame(frame1)

            # List all unique files in the folder
            file_paths = set(os.path.join(folder_path, f)
                             for f in os.listdir(folder_path)
                             if os.path.isfile(os.path.join(folder_path, f)))

            # Find the file input field
            file_input = WebDriverWait(self.driver, 30).until(expected_conditions.visibility_of_element_located(
                (By.XPATH, self.select_files_xpath)))
            print(f"{folder_path}")

            # Iterate over file paths and send each one individually
            for path in file_paths:
                file_input.send_keys(path + " ")
                time.sleep(10)
                # clear the file input filed
                WebDriverWait(self.driver, 10).until(expected_conditions.element_to_be_clickable(
                    (By.XPATH, self.clear_file_path))).click()

            WebDriverWait(self.driver, 60).until(expected_conditions.presence_of_element_located(
                (By.XPATH, self.complete_upload)))

        except Exception as e:
            raise e

    def store_file_names(self, path, folder_path):

        # Get a list of file names in the folder
        file_names = os.listdir(folder_path)

        # This will create a list of variables named file1, file2, file3, etc.
        variables = [file_name for file_name in file_names]
        # Write the values of the variables to the Excel file
        for i, data1 in enumerate(variables):
            substrings = data1.split("_")
            data = substrings[0]
            XLUtils.writeData(path, "Inputs", i + 2, 8, data)

    def validate_reg_documents(self, docnames):
        try:
            # Locate the all rows in document table
            rows = WebDriverWait(self.driver, 50).until(expected_conditions.visibility_of_element_located(
                (By.XPATH, "//div[@id='regDocTable']//table/tr")))

            for docname in docnames:
                regDoc = WebDriverWait(self.driver, 50).until(expected_conditions.presence_of_element_located(
                    (By.XPATH, "//a[contains(@href, '#') and text() = '" + docname + "']")))
                self.driver.save_screenshot(
                    ".\\Screenshots\\DocReg_screenshots\\docregistered.png")
                docno = regDoc.text
                if regDoc.is_displayed():
                    print(f"'{docno}' is displayed")
                    # assert docno == expected_docno, f"Expected text '{expected_docno}' but found '{docno}'"
                    assert True
                else:
                    print(f"'{docno}' is not displayed")
                    # assert docno == expected_docno, f"Expected text '{expected_docno}' but found '{docno}'"
                    self.driver.save_screenshot(
                        ".\\Screenshots\\DocReg_screenshots\\docNotRegistered.png")
                    assert False

        except Exception as e:
            raise e

    def open_document_edit_form(self):
        '''
                Method to wait till completion of file upload
                '''
        WebDriverWait(self.driver, 60).until(expected_conditions.presence_of_element_located(
            (By.XPATH, self.complete_upload)))
        WebDriverWait(self.driver, 10).until(expected_conditions.presence_of_element_located(
            (By.XPATH, self.bt_mass_edit_du))).click()
        WebDriverWait(self.driver, 10).until(expected_conditions.presence_of_element_located(
            (By.XPATH, self.bt_mass_edit_1))).click()
        WebDriverWait(self.driver, 10).until(expected_conditions.presence_of_element_located(
            (By.XPATH, self.mass_edit_form)))

    def update_document_title(self, title):
        try:
            ''' Entering text into title field '''
            title_field = WebDriverWait(self.driver, 10).until(expected_conditions.presence_of_element_located(
                (By.XPATH, self.text_ms_title_field)))
            title_field.click()
            time.sleep(2)
            title_field.send_keys(title)
        except Exception as e:
            raise e

    def update_document_workorder(self, wo_number):
        try:
            ''' selecting wo from dropdown '''
            ms_wo_selection = WebDriverWait(self.driver, 10).until(expected_conditions.presence_of_element_located(
                (By.XPATH, self.drp_ms_wo_field)))
            ''' Clicking on dropdown '''
            ms_wo_selection.click()
            time.sleep(2)
            ''' Waiting till visibility of all options from dropdown '''
            wo_drp = WebDriverWait(self.driver, 10).until(expected_conditions.visibility_of_element_located(
                (By.XPATH, self.ms_wo_drp)))

            ''' finding desired option from dropdown '''
            desired_wo = (wo_drp.find_element
                          (By.XPATH, "//div[contains(@class, 'v-list-item__title') and text() = '" + wo_number + "']"))
            ''' scrolling to desired option in dropdown '''
            ActionChains(self.driver).scroll_to_element(desired_wo).perform()
            desired_wo.click()
        except Exception as e:
            raise e

    def update_document_status(self, status):
        try:
            ''' selecting status from dropdown '''
            ms_wo_selection = WebDriverWait(self.driver, 10).until(expected_conditions.presence_of_element_located(
                (By.XPATH, self.drp_ms_status_field)))
            ms_wo_selection.click()
            status_drp = WebDriverWait(self.driver, 10).until(expected_conditions.visibility_of_element_located(
                (By.XPATH, self.ms_status_drp)))

            desired_status = (status_drp.find_element
                              (By.XPATH,
                               "// div[contains(@class , 'v-list-item__title') and text() = '" + status + "']"))
            ActionChains(self.driver).scroll_to_element(desired_status).perform()
            desired_status.click()
        except Exception as e:
            raise e

    def update_document_stage(self, stage):
        try:
            ''' selecting stage from dropdown '''
            ms_wo_selection = WebDriverWait(self.driver, 10).until(expected_conditions.presence_of_element_located(
                (By.XPATH, self.drp_ms_stage_field)))
            ms_wo_selection.click()
            stage_drp = WebDriverWait(self.driver, 10).until(expected_conditions.visibility_of_element_located(
                (By.XPATH, self.ms_stage_drp)))

            desired_stage = (stage_drp.find_element
                             (By.XPATH, "// div[contains(@class , 'v-list-item__title') and text() = '" + stage + "']"))
            ActionChains(self.driver).scroll_to_element(desired_stage).perform()
            desired_stage.click()
            time.sleep(3)
        except Exception as e:
            raise e

    def test_save_document(self):
        try:
            WebDriverWait(self.driver, 10).until(expected_conditions.element_to_be_clickable(
                (By.XPATH, self.bt_update_all))).click()

            WebDriverWait(self.driver, 10).until(expected_conditions.element_to_be_clickable(
                (By.XPATH, self.bt_test))).click()

            WebDriverWait(self.driver, 10).until(expected_conditions.element_to_be_clickable(
                (By.XPATH, self.bt_save))).click()
        except Exception as e:
            raise e

    def select_document(self, docname):
        try:
            # Locate the all rows in document table
            rows = WebDriverWait(self.driver, 50).until(expected_conditions.visibility_of_element_located(
                (By.XPATH, "//div[@id='regDocTable']//table/tr")))
            # # self.logger.info("table is identified")

            checkbox = (rows.find_element(By.XPATH,
                                          "//a[contains(@href, '#') and text() = '" + docname + "']/../..//div[@class='v-input--selection-controls__ripple']"))
            # # self.logger.info("Check box is identified")
            # Click on Checkbox
            ActionChains(self.driver).scroll_to_element(checkbox).perform()
            checkbox.click()
            # # self.logger.info("Check box is selected")

        except Exception as e:
            raise e

    def click_on_create_wf(self):
        try:
            WebDriverWait(self.driver, 10).until(expected_conditions.element_to_be_clickable(
                (By.XPATH, self.bt_plus_menu))).click()
            WebDriverWait(self.driver, 10).until(expected_conditions.element_to_be_clickable(
                (By.XPATH, self.bt_create_workflow))).click()
        except Exception as e:
            raise e

    def select_wf_route_template(self, wf_route_template):
        try:
            # select route template
            WebDriverWait(self.driver, 10).until(expected_conditions.element_to_be_clickable(
                (By.XPATH, self.wf_template_search))).click()
            # # self.logger.info("Clicked on search")
            self.driver.switch_to.default_content()
            WebDriverWait(self.driver, 100).until(expected_conditions.visibility_of_element_located(
                (By.XPATH, self.route_tmp_page)))
            # # self.logger.info("Route template page opened")
            try:
                template = WebDriverWait(self.driver, 100).until(expected_conditions.element_to_be_clickable(
                    (By.XPATH,
                     "//div[@class='wux-controls-responsivetileview-maincontent']//span[@class='searchItemSpan search_item_in_apps' and text() = '" + wf_route_template + "']")))
                template.click()
            except StaleElementReferenceException:
                template = WebDriverWait(self.driver, 100).until(expected_conditions.element_to_be_clickable(
                    (By.XPATH,
                     "//div[@class='wux-controls-responsivetileview-maincontent']//span[@class='searchItemSpan search_item_in_apps' and text() = '" + wf_route_template + "']")))
                template.click()
            # # self.logger.info("Route template selected")

            WebDriverWait(self.driver, 10).until(expected_conditions.element_to_be_clickable(
                (By.XPATH, self.bt_route_ok))).click()
            time.sleep(3)
        except Exception as e:
            raise e

    def create_wf(self, wf_title, reasonforissue):
        try:
            self.driver.switch_to.default_content()
            frame1 = WebDriverWait(self.driver, 30).until(expected_conditions.presence_of_element_located(
                (By.XPATH, self.iframeNMDocumentCntrl_xpath)))
            self.driver.switch_to.frame(frame1)
            WebDriverWait(self.driver, 60).until(expected_conditions.element_to_be_clickable(
                (By.XPATH,
                 "//select[@name='IMP_WFReasonForIssue']/option[contains(text(), '" + reasonforissue + "')]"))).click()
            WebDriverWait(self.driver, 10).until(expected_conditions.element_to_be_clickable(
                (By.XPATH, self.wf_title))).send_keys(wf_title)
            WebDriverWait(self.driver, 40).until(expected_conditions.element_to_be_clickable(
                (By.XPATH, self.bt_wf_create))).click()

        except Exception as e:
            raise e

    def start_WF(self, path):
        try:
            self.driver.switch_to.default_content()
            frame1 = WebDriverWait(self.driver, 100).until(expected_conditions.presence_of_element_located(
                (By.XPATH, self.iframerouteMg_xpath)))
            self.driver.switch_to.frame(frame1)
            wf = WebDriverWait(self.driver, 60).until(expected_conditions.visibility_of_element_located(
                (By.XPATH, self.created_wf)))
            wfname = wf.text

            # Store the WF name in excel input
            XLUtils.writeData(path, "Inputs", 2, 14, wfname)

            # # self.logger.info("Created WF is displayed" + wfname)
            ActionChains(self.driver).double_click(wf).perform()
            try:
                lnk_start = WebDriverWait(self.driver, 100).until(expected_conditions.element_to_be_clickable(
                    (By.XPATH, self.lnk_start_wf)))
                time.sleep(3)
                lnk_start.click()
                time.sleep(10)
                # # self.logger.info("Created WF is started")

            except Exception as e:
                raise e
        except Exception as e:
            raise e

    def validate_started_wf(self, expected_state):
        try:
            state = WebDriverWait(self.driver, 100).until(expected_conditions.visibility_of_element_located(
                (By.XPATH, self.awaiting_approval_xpath)))
            current_state = state.text
            if current_state == expected_state:
                print("Workflow is started")
            else:
                print("Workflow is not started")
            self.driver.save_screenshot(
                ".\\Screenshots\\workflow\\workflow_started.png")
            time.sleep(3)

        except Exception as e:
            raise e


class worklow_mg:
    # iframe_wf_widget = "//div[@class='module dock-item wi-9uao55t0g06q-siHcm1W leaf-widget moduleEditable']//iframe[starts-with(@id, 'frame')]"
    iframe_wf_widget = "//div[contains(@class,'IMP_WorkFlowManagment/IMP_WorkFlowManagment')]//iframe[starts-with(@id, 'frame')]"
    txt_wf_count = "//label[contains(., 'Number of WFs ')]"
    filter_xpath = "//span[@class='fa fa-filter dropbtn']"
    filter1_selected = "//div[@attr-selected='true' and @attr='filter' and text()='Completed']"
    filter1_notselected = "//div[@attr-selected='false' and @attr='filter' and text()='Completed']"
    filter2_selected = "//div[@attr-selected='true' and @attr='expand' and text()='Owned By Me']"
    filter2_notselected = "//div[@attr-selected='false' and @attr='expand' and text()='Owned By Me']"
    create_swf_icon = "//span[@title='Create SubWorkFlow']"
    swf_creation_page = "//span[@class='floatingPanel_title']/span[text()='Create Sub WorkFlow']"
    swf_Title = "//input[@name='IMP_WFTitle']"
    swf_template_search = "//div[(@class='fa fa-search search') and (@linked='routeTemplateOID')]"
    route_tmp_page = "//div[@class=' search-content-set-detail content-set-detail set-detail']"
    bt_ok = "//button[@id='id_in_app_ok' and text() = 'OK']"
    content_search = "//div[(@class='fa fa-search search') and (@linked='contentList')]"
    content_page = "//div[@class=' search-content-set-detail content-set-detail set-detail']"
    doc_page = "//div[@class='wux-layouts-gridengine-scroller-container']"
    create_btn = "//div[starts-with(@id,'formdivision')]//following-sibling::div/button[@class='btn-primary btn hd-form-submit-hover hd-form-submit-focus']"
    # pg_route_mg = "//div[@class='moduleHeader__title' and contains(text(), 'Route Management')]"
    iframerouteMg_xpath = "//div[@class='moduleWrapper']//iframe[starts-with(@id, 'frame-preview-')]"
    created_swf = "//div[@class='id-card-title-section']//span[contains(text(), 'SWF-')]"
    lnk_start_swf = "//div[@id='channel1']/div[@class='maturity-state-container']//a[contains(text(), 'Start')]"
    awaiting_comment_xpath = "//div[@id='channel1']/div[@class='maturity-state-container']//span[@class='rt-activity-state' and @title='Awaiting Comment']"
    create_cmt_icon = "//span[@title='Create Comment']"
    cmt_creation_page = "//span[@class='floatingPanel_title']/span[text()='Create Comment']"
    cmt_description = "//textarea[@name='description' and contains(@class,'form-control')]"
    iframe_crs_widget = "//div[contains(@class,'IMP_CRSManagment/IMP_CRSManagment')]//iframe[starts-with(@id, 'frame')]"
    crs_mg_widget = "//div[contains(@class,'moduleHeader__title') and text()='CRS Management']"
    bulk_modify_btn = "//span[@class='bulkmodifyCRS']"
    modify_form = "//span[contains(text(), 'Select Attribute:')]"
    select_attb = "//span[contains(text(), 'Select Attribute:')]/following-sibling::select[@class='form-control']"
    status_values_filed = "//div[@style='padding: 7px;']//option[text()='A - Work May Proceed']/.."
    bulk_modify_save_btn = "//span[contains(text(), 'Select Attribute:')]/..//following-sibling::div[contains(@style, 'display:')]//span[contains(text(), 'Save')]"
    textarea1 = "//span[contains(text(), 'Select Attribute:')]/..//following-sibling::div[@style='padding: 7px;']/textarea[@class='form-control']"
    textarea2 = "//span[contains(text(), 'Select Attribute:')]/..//following-sibling::div[@style='padding: 7px;']/input[@style='']"
    remove_btn = "//span[@class='removeCRS']"
    remove_confirm_form = "//div[contains(text(), 'Do you want to remove the selected CRS comments?')]"
    agree_btn = "//span[contains(text(), 'Agree')]"
    disagree_btn = "//span[contains(text(), 'Disagree')]"
    crs_filter = "//span[@class='fa fa-filter dropbtn']"
    # ur_status = "//tr[@type='Inbox Task']//div[@class='tddivcommon wrap' and @title='Abdallah Ajam']/ancestor::tr/following-sibling::tr[contains(@class,'impgrid')]//div[@value='UR']"
    remove_comment_success_msg = "//div[@role='alert' and @class='alert alert-success alert-dismissible' and contains(text(),'Object Marked As Removed')]"
    load_wf_page = "//div[starts-with(@title, 'WF')]"
    export_wf_icon = "//span[@title='Export Workflows to Excel']"
    export_crs_icon = "//span[@title='Export CRS']"
    crs_export_success_msg = "//div[@role='alert' and @class='alert alert-success alert-dismissible' and contains(text(),'File Successfully Downloaded')]"
    cust_column_filter_icon = "//span[@title='Customize Column Filter']"
    column_list = "//div[@class='dropdown-table-column-filter']/div"
    complete_task_icon = "//span[@title='Complete Task']"
    task_completion_success_msg = "//div[@role='alert' and @class='alert alert-success alert-dismissible' and contains(text(),'Task has been successfully completed')]"

    # constructer to initialise diver
    def __init__(self, driver):
        self.driver = driver

    # logger = LogGen.loggen()

    def select_WF_filter(self, filter1, filter2):
        '''
                Method to select the filter to see appropriate workflows
                '''
        try:
            self.driver.switch_to.default_content()
            frame1 = WebDriverWait(self.driver, 30).until(expected_conditions.presence_of_element_located(
                (By.XPATH, self.iframe_wf_widget)))
            self.driver.switch_to.frame(frame1)
            WebDriverWait(self.driver, 30).until(expected_conditions.presence_of_element_located(
                (By.XPATH, self.txt_wf_count)))
            time.sleep(30)
            # click on filter
            filterbt1 = WebDriverWait(self.driver, 30).until(expected_conditions.element_to_be_clickable(
                (By.XPATH, self.filter_xpath)))
            self.driver.execute_script("arguments[0].click();", filterbt1)
            time.sleep(3)

            try:
                # filter1 is already selected
                WebDriverWait(self.driver, 30).until(expected_conditions.presence_of_element_located(
                    (By.XPATH, "//div[@attr-selected='true' and @attr='filter' and text()='" + filter1 + "']")))
                # # self.logger.info("** Filter1 is already selected **")
                try:
                    # filter2 is already selected - select filter2
                    WebDriverWait(self.driver, 30).until(expected_conditions.element_to_be_clickable(
                        (By.XPATH,
                         "//div[@attr-selected='true' and @attr='expand' and text()='" + filter2 + "']"))).click()
                    # # self.logger.info("** filter2 is already selected select filter2**")
                except:
                    # filter2 is not selected - select filter2
                    WebDriverWait(self.driver, 30).until(expected_conditions.element_to_be_clickable(
                        (By.XPATH,
                         "//div[@attr-selected='false' and @attr='expand' and text()='" + filter2 + "']"))).click()
                    # # self.logger.info("** filter2 is not selected - select filter2 **")
                    time.sleep(10)

            except:
                # filter1 is not selected - select filter1
                filter1 = WebDriverWait(self.driver, 30).until(expected_conditions.element_to_be_clickable(
                    (By.XPATH, "//div[@attr-selected='false' and @attr='filter' and text()='" + filter1 + "']")))
                filter1.click()
                # # self.logger.info("** filter1 is not selected - select filter1 **")
                # select filter
                filterbt2 = WebDriverWait(self.driver, 30).until(expected_conditions.element_to_be_clickable(
                    (By.XPATH, self.filter_xpath)))
                self.driver.execute_script("arguments[0].click();", filterbt2)
                # # self.logger.info("** Filter1 is opened **")
                time.sleep(3)
                try:
                    # filter2 is already selected - select filter2
                    WebDriverWait(self.driver, 30).until(expected_conditions.element_to_be_clickable(
                        (By.XPATH, self.filter2_selected))).click()
                    # # self.logger.info("** filter2 is already selected - select filter2 **")
                except:
                    # filter2 is not selected - select filter2
                    WebDriverWait(self.driver, 30).until(expected_conditions.element_to_be_clickable(
                        (By.XPATH, self.filter2_notselected))).click()
                    # # self.logger.info("** filter2 is not selected - select filter2 **")
        except Exception as e:
            raise e

    def select_Workflow(self, wf):
        '''
                Method to select the desired workflow
                '''
        try:
            self.driver.switch_to.default_content()
            frame1 = WebDriverWait(self.driver, 100).until(expected_conditions.presence_of_element_located(
                (By.XPATH, self.iframe_wf_widget)))
            self.driver.switch_to.frame(frame1)
            WebDriverWait(self.driver, 100).until(expected_conditions.element_to_be_clickable(
                (By.XPATH, "//div[@title='" + wf + "']"))).click()
            time.sleep(10)
            self.driver.save_screenshot(
                ".\\Screenshots\\subworkflow\\wf_selected.png")
        except Exception as e:
            raise e

    def select_inboxTask(self, assignee):
        try:
            WebDriverWait(self.driver, 30).until(expected_conditions.element_to_be_clickable(
                (By.XPATH,
                 "//tr[@type='Inbox Task']//div[@class='tddivcommon wrap' and @title='" + assignee + "']/ancestor::tr[@class='impgrid TRDetails expanded branch']"))).click()
        except Exception as e:
            raise e

    def click_on_create_swf(self):
        try:
            WebDriverWait(self.driver, 30).until(expected_conditions.element_to_be_clickable(
                (By.XPATH, self.create_swf_icon))).click()
            WebDriverWait(self.driver, 30).until(expected_conditions.element_to_be_clickable(
                (By.XPATH, self.swf_creation_page)))
            time.sleep(5)
            self.driver.save_screenshot(
                ".\\Screenshots\\subworkflow\\create_subworkflow_form.png")

        except Exception as e:
            raise e

    def select_swf_route_template(self, swf_route_template):
        try:
            # select route template
            WebDriverWait(self.driver, 10).until(expected_conditions.element_to_be_clickable(
                (By.XPATH, self.swf_template_search))).click()
            # # self.logger.info("Clicked on search")
            self.driver.switch_to.default_content()
            WebDriverWait(self.driver, 100).until(expected_conditions.visibility_of_element_located(
                (By.XPATH, self.route_tmp_page)))
            # # self.logger.info("Route template page opened")
            try:
                template = WebDriverWait(self.driver, 100).until(expected_conditions.element_to_be_clickable(
                    (By.XPATH,
                     "//div[@class='wux-controls-responsivetileview-maincontent']//span[@class='searchItemSpan search_item_in_apps' and text() = '" + swf_route_template + "']")))
                template.click()
            except StaleElementReferenceException:
                template = WebDriverWait(self.driver, 100).until(expected_conditions.element_to_be_clickable(
                    (By.XPATH,
                     "//div[@class='wux-controls-responsivetileview-maincontent']//span[@class='searchItemSpan search_item_in_apps' and text() = '" + swf_route_template + "']")))
                template.click()

            # # self.logger.info("Route template selected")
            WebDriverWait(self.driver, 10).until(expected_conditions.element_to_be_clickable(
                (By.XPATH, self.bt_ok))).click()
        except Exception as e:
            raise e

    # select single or multiple content
    def select_content_s_or_m(self, contents):
        try:
            # Click on content search
            self.driver.switch_to.default_content()
            frame1 = WebDriverWait(self.driver, 30).until(expected_conditions.presence_of_element_located(
                (By.XPATH, self.iframe_wf_widget)))
            self.driver.switch_to.frame(frame1)
            WebDriverWait(self.driver, 10).until(expected_conditions.element_to_be_clickable(
                (By.XPATH, self.content_search))).click()

            self.driver.switch_to.default_content()
            WebDriverWait(self.driver, 100).until(expected_conditions.visibility_of_element_located(
                (By.XPATH, self.content_page)))

            actions = ActionChains(self.driver)
            for doc in contents:
                # # self.logger.info("** control key pressed **")
                content = WebDriverWait(self.driver, 100).until(expected_conditions.element_to_be_clickable(
                    (By.XPATH,
                     "//span[@class='searchItemSpan' and text() = '" + doc + "']/ancestor::div[@class='wux-controls-responsivetileview-maincontent']")))
                content.click()
                # # self.logger.info("** doc selected **")
                time.sleep(2)
                # Press and Hold the Ctrl key
                actions.key_down(Keys.CONTROL)
            # Release control key
            actions.key_up(Keys.CONTROL)
            # # self.logger.info("** control key released **")
            actions.perform()
            # # self.logger.info("content selected")
            WebDriverWait(self.driver, 10).until(expected_conditions.element_to_be_clickable(
                (By.XPATH, self.bt_ok))).click()

        except Exception as e:
            raise e

    # Select all content
    def select_content_all(self, doc1):
        try:
            # Click on content search
            self.driver.switch_to.default_content()
            frame1 = WebDriverWait(self.driver, 30).until(expected_conditions.presence_of_element_located(
                (By.XPATH, self.iframe_wf_widget)))
            self.driver.switch_to.frame(frame1)
            time.sleep(2)
            WebDriverWait(self.driver, 60).until(expected_conditions.element_to_be_clickable(
                (By.XPATH, self.content_search))).click()
            time.sleep(2)

            self.driver.switch_to.default_content()
            WebDriverWait(self.driver, 100).until(expected_conditions.visibility_of_element_located(
                (By.XPATH, self.content_page)))
            try:
                content = WebDriverWait(self.driver, 100).until(expected_conditions.visibility_of_element_located(
                    (By.XPATH,
                     "//span[@class='searchItemSpan' and text() = '" + doc1 + "']/ancestor::div[@class='wux-controls-responsivetileview-maincontent']")))
                content.click()
            except StaleElementReferenceException:
                content = WebDriverWait(self.driver, 100).until(expected_conditions.visibility_of_element_located(
                    (By.XPATH,
                     "//span[@class='searchItemSpan' and text() = '" + doc1 + "']/ancestor::div[@class='wux-controls-responsivetileview-maincontent']")))
                content.click()

            actions = ActionChains(self.driver)
            # Press and Hold the Ctrl key
            actions.key_down(Keys.CONTROL)
            time.sleep(2)
            actions.send_keys('a')
            time.sleep(1)
            actions.key_up('a')
            # Release the Ctrl key
            actions.key_up(Keys.CONTROL)
            # # self.logger.info("** control key released **")
            actions.perform()
            # # self.logger.info("** action performed **")
            time.sleep(2)
            # # self.logger.info("content selected")
            WebDriverWait(self.driver, 10).until(expected_conditions.element_to_be_clickable(
                (By.XPATH, self.bt_ok))).click()

        except Exception as e:
            raise e

    def create_swf(self, swf_title):
        '''
                Method to create the subworkflow
                '''
        try:
            self.driver.switch_to.default_content()
            frame1 = WebDriverWait(self.driver, 100).until(expected_conditions.presence_of_element_located(
                (By.XPATH, self.iframe_wf_widget)))
            self.driver.switch_to.frame(frame1)
            title = WebDriverWait(self.driver, 100).until(expected_conditions.element_to_be_clickable(
                (By.XPATH, self.swf_Title)))
            title.send_keys(swf_title)
            time.sleep(3)
            WebDriverWait(self.driver, 60).until(expected_conditions.presence_of_element_located(
                (By.XPATH, self.create_btn)))
            WebDriverWait(self.driver, 60).until(expected_conditions.element_to_be_clickable(
                (By.XPATH, self.create_btn))).click()
            time.sleep(3)
            self.driver.save_screenshot(
                ".\\Screenshots\\subworkflow\\swf_created.png")
        except Exception as e:
            raise e

    def start_SWF(self, path):
        try:
            # WebDriverWait(self.driver, 150).until(expected_conditions.visibility_of_element_located(
            #     (By.XPATH, self.pg_route_mg)))
            time.sleep(2)
            self.driver.switch_to.default_content()
            frame1 = WebDriverWait(self.driver, 100).until(expected_conditions.presence_of_element_located(
                (By.XPATH, self.iframerouteMg_xpath)))
            self.driver.switch_to.frame(frame1)

            swf = WebDriverWait(self.driver, 60).until(expected_conditions.visibility_of_element_located(
                (By.XPATH, self.created_swf)))
            swfname = swf.text

            # Store the SWF name in excel input
            XLUtils.writeData(path, "Inputs", 3, 14, swfname)
            try:
                lnk_start = WebDriverWait(self.driver, 100).until(expected_conditions.element_to_be_clickable(
                    (By.XPATH, self.lnk_start_swf)))
                time.sleep(3)
                lnk_start.click()
                time.sleep(10)
                # # self.logger.info("Created SWF is started")

            except Exception as e:
                raise e
        except Exception as e:
            raise e

    def validate_started_swf(self, expected_state):
        try:
            state = WebDriverWait(self.driver, 100).until(expected_conditions.visibility_of_element_located(
                (By.XPATH, self.awaiting_comment_xpath)))
            current_state = state.text
            if current_state == expected_state:
                print("SubWorkflow is started")
            else:
                print("SubWorkflow is not started")
            self.driver.save_screenshot(
                ".\\Screenshots\\subworkflow\\Subworkflow_started.png")
            time.sleep(3)

        except Exception as e:
            raise e

    def expand_Workflow(self, wf):
        '''
                Method to expand the desired workflow
                '''
        try:
            self.driver.switch_to.default_content()
            frame1 = WebDriverWait(self.driver, 100).until(expected_conditions.presence_of_element_located(
                (By.XPATH, self.iframe_wf_widget)))
            self.driver.switch_to.frame(frame1)
            WebDriverWait(self.driver, 100).until(expected_conditions.presence_of_element_located(
                (By.XPATH, "//div[@title='" + wf + "']")))
            WebDriverWait(self.driver, 100).until(expected_conditions.element_to_be_clickable(
                (By.XPATH, "//div[@title='" + wf + "']/parent::*/preceding-sibling::div[@class='expand']"))).click()
            time.sleep(10)
            self.driver.save_screenshot(
                ".\\Screenshots\\Comment_Creation\\wf_expanded.png")
        except Exception as e:
            raise e

    def select_Subworkflow(self, swf):
        '''
                Method to select the desired subworkflow
                '''
        try:
            self.driver.switch_to.default_content()
            frame1 = WebDriverWait(self.driver, 100).until(expected_conditions.presence_of_element_located(
                (By.XPATH, self.iframe_wf_widget)))
            self.driver.switch_to.frame(frame1)
            WebDriverWait(self.driver, 100).until(expected_conditions.element_to_be_clickable(
                (By.XPATH, "//div[@title='" + swf + "']"))).click()
            time.sleep(10)
            self.driver.save_screenshot(
                ".\\Screenshots\\Comment_Creation\\swf_selected.png")
        except Exception as e:
            raise e

    def select_doc_under_task(self, assignee, doc):
        try:
            WebDriverWait(self.driver, 30).until(expected_conditions.element_to_be_clickable(
                (By.XPATH,
                 "//tr[@type='Inbox Task']//div[@class='tddivcommon wrap' and @title='" + assignee + "']/ancestor::tr[@class='impgrid TRDetails expanded branch']")))
            WebDriverWait(self.driver, 30).until(expected_conditions.element_to_be_clickable(
                (By.XPATH,
                 "//tr[@type='Inbox Task']//div[@class='tddivcommon wrap' and @title='" + assignee + "']/ancestor::tr[@class='impgrid TRDetails expanded branch']/following-sibling::tr//div[@class='td-first-div wrap' and @title='" + doc + "']"))).click()
        except Exception as e:
            raise e

    def click_on_create_comment(self):
        try:
            WebDriverWait(self.driver, 30).until(expected_conditions.element_to_be_clickable(
                (By.XPATH, self.create_cmt_icon))).click()
            WebDriverWait(self.driver, 30).until(expected_conditions.element_to_be_clickable(
                (By.XPATH, self.cmt_creation_page)))
            time.sleep(5)
            self.driver.save_screenshot(
                ".\\Screenshots\\Comment_Creation\\create_cmt_form.png")

        except Exception as e:
            raise e

    def create_cmt(self, desc, status):
        '''
                Method to create comment
                '''
        try:
            # self.driver.switch_to.default_content()
            # frame1 = WebDriverWait(self.driver, 100).until(expected_conditions.presence_of_element_located(
            #     (By.XPATH, self.iframe_wf_widget)))
            # self.driver.switch_to.frame(frame1)
            description = WebDriverWait(self.driver, 100).until(expected_conditions.element_to_be_clickable(
                (By.XPATH, self.cmt_description)))
            description.send_keys(desc)
            select_status = WebDriverWait(self.driver, 100).until(expected_conditions.presence_of_element_located(
                (By.XPATH,
                 "//select[@type='select' and @name='attribute[IMP_CRSInitialStatus]']/option[@value='" + status + "']")))
            select_status.click()
            WebDriverWait(self.driver, 30).until(expected_conditions.element_to_be_clickable(
                (By.XPATH, self.create_btn))).click()
            time.sleep(7)
            self.driver.save_screenshot(
                ".\\Screenshots\\Comment_Creation\\Comment_created.png")
        except Exception as e:
            raise e

    # def validate_doc_status(self,expected_statuses,assignee,doc):
    #     try:
    #         # status = WebDriverWait(self.driver, 100).until(expected_conditions.visibility_of_element_located(
    #         #     (By.XPATH, "//div[@class='tddivcommon wrap' and @title='D - Rejected']")))
    #         status = WebDriverWait(self.driver, 30).until(expected_conditions.element_to_be_clickable(
    #             (By.XPATH,
    #              "//tr[@type='Inbox Task']//div[@class='tddivcommon wrap' and @title='" + assignee + "']"
    #             "/ancestor::tr[@class='impgrid TRDetails expanded branch']"
    #             "/following-sibling::tr//div[@class='td-first-div wrap' and @title='" + doc + "']/ancestor::tr[@class='impgrid TRDetails expanded branch']//td[@class='td-common' and @name='IMP_WFReviewOutcome'")))
    #         current_status = status.text
    #         if current_status in expected_statuses:
    #             print("Doc Status is updated")
    #         else:
    #             print("Doc Status is not updated to any of the expected statuses: {', '.join(expected_statuses)}")
    #         self.driver.save_screenshot(".\\Screenshots\\Comment_Creation\\doc status.png")
    #         time.sleep(3)
    #
    #     except Exception as e:
    #         raise e

    def navigate_to_CRS_mg(self):
        '''
                Method to navigate to CRS Management
                '''
        try:
            self.driver.switch_to.default_content()
            crs_widget = WebDriverWait(self.driver, 100).until(expected_conditions.visibility_of_element_located(
                (By.XPATH, self.crs_mg_widget)))
            # Assertion for the presence of CRS widget
            assert crs_widget.is_displayed(), "CRS Management widget is not displayed"

            # self.driver.switch_to.default_content()
            frame1 = WebDriverWait(self.driver, 100).until(expected_conditions.presence_of_element_located(
                (By.XPATH, self.iframe_crs_widget)))
            self.driver.switch_to.frame(frame1)
            # # self.logger.info("** navigated to CRS Management widget **")

        except Exception as e:
            raise e

    def validate_created_cmt(self, desc):
        '''
                        Method to validate created comment on CRS widget by description
                        '''
        try:
            descrip = WebDriverWait(self.driver, 100).until(expected_conditions.visibility_of_element_located(
                (By.XPATH, "//div[contains(@class,'description wrap') and contains(text(), '" + desc + "')]")))

            self.driver.save_screenshot(
                ".\\Screenshots\\Comment_Creation\\comment_created.png")
            # Assertion for the presence of the description
            assert descrip.is_displayed(), f"Description '{desc}' is not displayed"
        except Exception as e:
            raise e

    def edit_single_comment(self, desc, desc_update):
        '''
                        Method to edit single comment by double-click on it
                        '''
        try:
            comment1 = WebDriverWait(self.driver, 100).until(expected_conditions.visibility_of_element_located(
                (By.XPATH, "//div[contains(@class,'description wrap') and contains(text(), '" + desc + "')]")))
            WebDriverWait(self.driver, 100).until(expected_conditions.element_to_be_clickable(
                (By.XPATH, "//div[contains(@class,'description wrap') and contains(text(), '" + desc + "')]")))
            ActionChains(self.driver).double_click(comment1).perform()
            # # self.logger.info("*** double clicked on desired comment ***")
            WebDriverWait(self.driver, 100).until(expected_conditions.visibility_of_element_located(
                (By.XPATH, "//div[@class='popup']")))
            # # self.logger.info("*** comment form is opened ***")
            description = WebDriverWait(self.driver, 100).until(expected_conditions.visibility_of_element_located(
                (By.XPATH, "//textarea[@class='form-control' and @name='description']")))
            description.clear()
            description.send_keys(desc_update)
            WebDriverWait(self.driver, 100).until(expected_conditions.visibility_of_element_located(
                (By.XPATH, "//button[@type='button' and text()='Update']"))).click()
            time.sleep(6)
            self.driver.save_screenshot(
                ".\\Screenshots\\Comment_Creation\\comment_updated.png")
        except Exception as e:
            raise e

    def modify_bulk_comments(self, index_list, attribute, status_value=None, comment=None):
        '''
                        Method to edit bulk comments by bulk modify CRS
                        '''
        try:
            # Press and hold the control key
            ActionChains(self.driver).key_down(Keys.CONTROL).perform()

            # Select multiple comments based on provided indexes
            for index in index_list:
                comments = WebDriverWait(self.driver, 100).until(
                    expected_conditions.element_to_be_clickable(
                        (By.XPATH, "(//tr[@class='impgrid TRDetails'])[" + str(index) + "]")))
                comments.click()

            # Release the control key
            ActionChains(self.driver).key_up(Keys.CONTROL).perform()
            # # self.logger.info("*** Desired comments are selected***")

            WebDriverWait(self.driver, 20).until(expected_conditions.element_to_be_clickable(
                (By.XPATH, self.bulk_modify_btn))).click()

            WebDriverWait(self.driver, 40).until(expected_conditions.visibility_of_element_located(
                (By.XPATH, self.modify_form)))
            # # self.logger.info("*** comment form is opened ***")
            try:
                # select attribute path
                select_element = WebDriverWait(self.driver, 10).until(expected_conditions.element_to_be_clickable(
                    (By.XPATH, self.select_attb)))

                select_element.click()

                # select desired attribute based on given input
                WebDriverWait(self.driver, 10).until(expected_conditions.element_to_be_clickable(
                    (By.XPATH,
                     "//span[contains(text(), 'Select Attribute:')]/following-sibling::select[@class='form-control']/option[text()='" + attribute + "']"))).click()

                # # self.logger.info("*** Desired attribute selected ***")

                select = Select(select_element)
                selected_option_text = select.first_selected_option.text
                print("Selected option is:", selected_option_text)

                # save button path
                save = (WebDriverWait(self.driver, 10).until(expected_conditions.element_to_be_clickable(
                    (By.XPATH, self.bulk_modify_save_btn))))

            except Exception as e:
                raise e
            try:
                if selected_option_text == 'Initial Status':

                    dropdown = (WebDriverWait(self.driver, 10).until(expected_conditions.element_to_be_clickable(
                        (By.XPATH, self.status_values_filed))))
                    dropdown.click()
                    option1 = (WebDriverWait(self.driver, 10).until(expected_conditions.element_to_be_clickable(
                        (By.XPATH,
                         "//div[@style='padding: 7px;']//option[@value='" + status_value + "']"))))
                    option1.click()
                    save.click()

                elif selected_option_text == 'Design Reviewer/PMC Comment':
                    textarea = (WebDriverWait(self.driver, 10).until(expected_conditions.visibility_of_element_located(
                        (By.XPATH, self.textarea1))))
                    textarea.send_keys(comment)

                    print("comment modified ")
                    save.click()

                elif selected_option_text == 'Doc/Section/Page':
                    textarea = (WebDriverWait(self.driver, 10).until(expected_conditions.visibility_of_element_located(
                        (By.XPATH, self.textarea2))))
                    print("text area identified ")
                    textarea.click()
                    print("clicked on text area")
                    textarea.send_keys(comment)

                    print("Doc/Section/Page modified ")
                    save.click()
            except Exception as e:
                raise e
        except Exception as e:
            raise e

    def remove_comments(self, index_list):
        '''
                        Method to remove comments
                        '''
        try:

            # Press and hold the control key
            ActionChains(self.driver).key_down(Keys.CONTROL).perform()

            # Select multiple comments based on provided indexes
            for index in index_list:
                comments = WebDriverWait(self.driver, 100).until(
                    expected_conditions.element_to_be_clickable(
                        (By.XPATH, "(//tr[@class='impgrid TRDetails'])[" + str(index) + "]")))
                comments.click()

            # Release the control key
            ActionChains(self.driver).key_up(Keys.CONTROL).perform()
            # # self.logger.info("*** Desired comments are selected***")

            # Click on Remove button
            WebDriverWait(self.driver, 20).until(expected_conditions.element_to_be_clickable(
                (By.XPATH, self.remove_btn))).click()

            WebDriverWait(self.driver, 40).until(expected_conditions.visibility_of_element_located(
                (By.XPATH, self.remove_confirm_form)))
            # # self.logger.info("*** remove form is opened ***")
            try:
                # Click on AGREE button
                select_element = WebDriverWait(self.driver, 10).until(expected_conditions.element_to_be_clickable(
                    (By.XPATH, self.agree_btn)))

                select_element.click()

                remove_comment_success_message = WebDriverWait(self.driver, 40).until(
                    expected_conditions.visibility_of_element_located(
                        (By.XPATH, self.remove_comment_success_msg)))

                self.driver.save_screenshot(
                    ".\\Screenshots\\Comment_Modification\\comment_removed.png")

                # Assert that the success message is displayed
                assert remove_comment_success_message.is_displayed(), f"Success message is not displayed"

            except Exception as e:
                raise e

        except Exception as e:
            raise e

    def filtered_comments(self, filter):
        '''
                        Method to apply filter in CRS widgets
                                '''
        try:
            CRS_filter = WebDriverWait(self.driver, 10).until(expected_conditions.element_to_be_clickable(
                (By.XPATH, self.crs_filter)))
            CRS_filter.click()
            time.sleep(6)
            filter_value = WebDriverWait(self.driver, 10).until(expected_conditions.element_to_be_clickable(
                (By.XPATH, "//div[@class='dropdown-content']/div[contains(text(), '" + filter + "')]")))
            filter_value.click()
            time.sleep(10)
            self.driver.save_screenshot(
                ".\\Screenshots\\Comment_Modification\\Filtered_Comments.png")

        except Exception as e:
            raise e

    def export_Workflows(self):
        '''
                Method to export the workflow in the excel
                '''
        try:
            self.driver.switch_to.default_content()
            frame1 = WebDriverWait(self.driver, 100).until(expected_conditions.presence_of_element_located(
                (By.XPATH, self.iframe_wf_widget)))
            self.driver.switch_to.frame(frame1)
            # self.logger.info("** entered into frame**")

            # wait till workflow page loaded
            WebDriverWait(self.driver, 100).until(expected_conditions.element_to_be_clickable(
                (By.XPATH, self.load_wf_page)))
            # self.logger.info("** widget loaded**")
            # export workflows to excel icon element
            WebDriverWait(self.driver, 100).until(expected_conditions.element_to_be_clickable(
                (By.XPATH, self.export_wf_icon))).click()
            time.sleep(5)

        except Exception as e:
            raise e

    def export_CRS(self):
        '''
                Method to export the CRS report from WF
                '''
        try:
            # export CRS icon element
            export_CRS = WebDriverWait(self.driver, 100).until(expected_conditions.element_to_be_clickable(
                (By.XPATH, self.export_crs_icon)))
            time.sleep(3)
            export_CRS.click()
            # self.logger.info("** clicked on export CRS**")
            crs_exported_success_message = WebDriverWait(self.driver, 40).until(
                expected_conditions.visibility_of_element_located(
                    (By.XPATH, self.crs_export_success_msg)))

            # Assert that the success message is displayed
            assert crs_exported_success_message.is_displayed(), "Success message not displayed after exporting CRS"
            self.driver.save_screenshot(
                ".\\Screenshots\\workflow\\CRS_exported.png")
            time.sleep(3)
        except Exception as e:
            self.driver.save_screenshot(
                ".\\Screenshots\\workflow\\CRS_export_failure.png")
            raise e

    def click_on_customize_column_filter(self):
        '''
                Method to click on customize column filter
                '''
        try:
            time.sleep(2)
            WebDriverWait(self.driver, 100).until(expected_conditions.element_to_be_clickable(
                (By.XPATH, self.cust_column_filter_icon))).click()
            # self.logger.info("** clicked on customize column filter **")

        except Exception as e:
            raise e

    def customize_column_filter(self, columns):
        try:
            # Locate the list of columns
            list = WebDriverWait(self.driver, 30).until(expected_conditions.visibility_of_element_located(
                (By.XPATH, self.column_list)))

            for column in columns:
                colm = (list.find_element(By.XPATH,
                                          "//div[@class='dropdown-table-column-filter']/div[text()='" + column + "']"))

                self.driver.execute_script("arguments[0].scrollIntoView();", colm)
                print(f"** {column} is located **")

                # click on column
                try:
                    element = WebDriverWait(self.driver, 30).until(expected_conditions.visibility_of_element_located(
                        (By.XPATH,
                         "//div[@class='dropdown-table-column-filter']/div[text()='" + column + "'and @attr-selected]")))
                    element.click()
                    print(f"** {column} is unselected **")
                    time.sleep(2)

                except:
                    print(f"** {column} is already unselected **")
                    pass

                # click on column
                # try:
                #     element = WebDriverWait(self.driver, 30).until(expected_conditions.visibility_of_element_located(
                #     (By.XPATH, "//div[@class='dropdown-table-column-filter']/div[text()='" + column + "'and not(@attr-selected)]")))
                #     element.click()
                #     print(f"** {column} is selected **")
                #     time.sleep(2)
                #
                # except:
                #     print(f"** {column} is already selected **")
                #     pass

            time.sleep(5)
            self.driver.save_screenshot(
                ".\\Screenshots\\workflow\\column_customized.png")

        except Exception as e:
            raise e

    def complete_task(self, assignee):
        try:
            # Wait until the UR status element is visible
            element = WebDriverWait(self.driver, 30).until(expected_conditions.visibility_of_element_located(
                (By.XPATH,
                 "//tr[@type='Inbox Task']//div[@class='tddivcommon wrap' and @title='" + assignee + "']/ancestor::tr/following-sibling::tr[contains(@class,'impgrid')]//div[@value='UR']")))

            print("** Can not complete task as one of the document has UR status **")
            # self.logger.info("** Can not complete task as one of the document has UR status **")

        except:
            print("all documents has other status than UR ")
            try:
                # complete task element
                WebDriverWait(self.driver, 30).until(expected_conditions.element_to_be_clickable(
                    (By.XPATH, self.complete_task_icon))).click()

                task_completion_success_message = WebDriverWait(self.driver, 40).until(
                    expected_conditions.visibility_of_element_located(
                        (By.XPATH, self.task_completion_success_msg)))

                self.driver.save_screenshot(
                    ".\\Screenshots\\Task_completed.png")

                # Assert that the success message is displayed
                assert task_completion_success_message.is_displayed(), f"Success message is not displayed"

                # self.logger.info("** Task Completed **")
            except Exception as e:
                raise e


class import_comments:
    export_crs_template = "//span[@title='Export CRS Template']"
    select_review_entity_wdw = "//div[@class='menu']//span[text()='Select Review Entity:']"
    choose_items = "//div[@class='menu']//span[text()='Choose items']"
    close_items = "//div[@style='display: inline-grid; grid-auto-flow: column;']"
    disp_list = "//div[@class='MultiSelectList']/label"
    save_btn = "//span[text()='Select Review Entity:']/../following-sibling::div/button[@class='btn-primary btn hd-form-submit-hover hd-form-submit-focus']/span[text()='Save']"
    import_comments = "//span[@title='Import Comments']"
    import_comment_popup = "//div[@class='top']/div[@id='importcomment']"
    browse_crs_file = "//input[@id='inpSWFImportCRS']"
    browse_pdf_file = "//input[@id='inpSWFImportPDF']"
    clear_file_path = "//button[@class='v-icon notranslate v-icon--link mdi mdi-close theme--light']"
    import_save_btn = "//div[@class='top']/div[@id='importcomment']/..//following-sibling::div/button[@class='btn-primary btn hd-form-submit-hover hd-form-submit-focus']/span[text()='Save']"

    # constructer to initialise diver
    def __init__(self, driver):
        self.driver = driver

    # # logger = LogGen.loggen()
    def click_on_export_CRS_Template(self):
        try:
            WebDriverWait(self.driver, 30).until(expected_conditions.element_to_be_clickable(
                (By.XPATH, self.export_crs_template))).click()
            # # self.logger.info("*** Clicked on export CRS template button ***")
            WebDriverWait(self.driver, 30).until(expected_conditions.visibility_of_element_located(
                (By.XPATH, self.select_review_entity_wdw)))
            # # self.logger.info("*** select review entity window is opened ***")
            ch_item = WebDriverWait(self.driver, 30).until(expected_conditions.element_to_be_clickable(
                (By.XPATH, self.choose_items)))
            ch_item.click()
        except Exception as e:
            raise e

    def select_disp(self, disciplines):
        try:
            # Locate the list of disciplines
            list = WebDriverWait(self.driver, 30).until(expected_conditions.visibility_of_element_located(
                (By.XPATH, self.disp_list)))

            for discipline in disciplines:
                discp = (list.find_element(By.XPATH,
                                           "//div[@class='MultiSelectList']//span[contains(text(), '" + discipline + "')]"))

                self.driver.execute_script("arguments[0].scrollIntoView();", discp)
                # select the checkbox
                checkbox = WebDriverWait(self.driver, 30).until(expected_conditions.element_to_be_clickable(
                    (By.XPATH,
                     "//div[@class='MultiSelectList']//span[contains(text(), '" + discipline + "')]/ancestor::label/input[@type='checkbox']")))
                checkbox.click()
                time.sleep(5)
                self.driver.save_screenshot(
                    ".\\Screenshots\\Comment_Creation\\checkbox_selected.png")

        except Exception as e:
            raise e

    def click_on_save(self):
        try:
            WebDriverWait(self.driver, 30).until(expected_conditions.element_to_be_clickable(
                (By.XPATH, self.close_items))).click()
            WebDriverWait(self.driver, 30).until(expected_conditions.element_to_be_clickable(
                (By.XPATH, self.save_btn))).click()
            time.sleep(1)
            self.driver.save_screenshot(
                ".\\Screenshots\\Comment_Creation\\downloading_file.png")
            # # self.logger.info("*** File is downloaded ***")
        except Exception as e:
            raise e

    def click_on_Import_Comments(self):
        try:
            WebDriverWait(self.driver, 30).until(expected_conditions.element_to_be_clickable(
                (By.XPATH, self.import_comments))).click()
            # # self.logger.info("*** Clicked on Import comments button ***")
            WebDriverWait(self.driver, 30).until(expected_conditions.visibility_of_element_located(
                (By.XPATH, self.import_comment_popup)))
            # # self.logger.info("*** Import comments popup opened ***")
        except Exception as e:
            raise e

    def browse_CRS_file(self, crs_file_path):
        try:
            file_input = WebDriverWait(self.driver, 10).until(expected_conditions.presence_of_element_located(
                (By.XPATH, self.browse_crs_file)))
            # # self.logger.info("*** Clicked on browse ***")
            file_input.send_keys(crs_file_path)
            # # self.logger.info("*** CRS file is selected ***")
        except Exception as e:
            raise e

    def browse_annotation_file(self, pdf_file_path):
        try:
            pdf_input = WebDriverWait(self.driver, 10).until(expected_conditions.presence_of_element_located(
                (By.XPATH, self.browse_pdf_file)))
            # # self.logger.info("*** Clicked on browse annotations ***")
            pdf_input.send_keys(pdf_file_path)
            # # self.logger.info("*** annotation file is selected ***")
        except Exception as e:
            raise e

    def click_on_Save_Import_Comments(self):
        try:
            WebDriverWait(self.driver, 30).until(expected_conditions.element_to_be_clickable(
                (By.XPATH, self.import_save_btn))).click()
            # # self.logger.info("*** Clicked on Import save button ***")
            time.sleep(15)
            self.driver.save_screenshot(
                ".\\Screenshots\\Comment_Creation\\importing_comments.png")
        except Exception as e:
            raise e

    def validate_CRS_Comments_icon(self, assignee):
        try:
            CRS_icon = WebDriverWait(self.driver, 30).until(expected_conditions.presence_of_element_located(
                (By.XPATH,
                 "//tr[@type='Inbox Task']//div[@class='tddivcommon wrap' and @title='" + assignee + "']/ancestor::tr//div[@title='Click to Download File']")))

            # Scroll to CRS icon element
            actions = ActionChains(self.driver)
            actions.move_to_element(CRS_icon).perform()
            self.driver.execute_script("arguments[0].scrollIntoView();", CRS_icon)
            time.sleep(1)  # Optional: Wait for smooth scrolling effect
            # # self.logger.info("*** Clicked on Import save button ***")
            time.sleep(15)
            # Assertion for the presence of CRS icon
            assert CRS_icon.is_displayed(), "CRS icon is not displayed"
            self.driver.save_screenshot(
                ".\\Screenshots\\Comment_Creation\\comments_imported.png")
            # # self.logger.info("CRS icon is displayed successfully")
        except Exception as e:
            raise e


class regDocument_properties1:
    reg_rows_per_page_filter = "//div[@class='v-input__slot' and @aria-owns='list-119']//div[@class='v-select__slot']//div[@class='v-select__selection v-select__selection--comma' and text()='50']"
    # reg_rows_per_page_filter = "//div[@class='v-data-footer__icons-after']//button[@class='v-btn v-btn--icon v-btn--round v-btn--text theme--light v-size--default']"
    reg_rows_per_page_all = "//div[@class='v-list-item__content']//div[@class='v-list-item__title' and text()='All']"
    filter_xpath = "(//div[@class='v-input--selection-controls__ripple'])[3]"
    restore_xpath = "//div[contains(@class, 'moduleWrapper clearfix ifwe-tabview')]//div[@class='wp-tabview-panel']//div[@class='widget-dd-menu dropdown-menu dropdown-menu-root dropdown dropdown-root']//child::span[@class='maximize-icon fonticon fonticon-resize-small']"
    all_rows_selected = "//div[@class='v-input__slot' and @aria-owns='list-119']//div[@class='v-select__slot']//div[@class='v-select__selection v-select__selection--comma' and text()='All']"
    # document name panel - start
    doc_download_panel = "//div[@class='v-slide-group__content v-tabs-bar__content']//i[@class='v-icon notranslate v-icon--left mdi mdi-account-clock theme--light']"
    doc_transmittal_panel = "//div[@class='v-slide-group__content v-tabs-bar__content']//i[@class='v-icon notranslate v-icon--left mdi mdi-axis-arrow-lock theme--light']"
    doc_wf_panel = "//div[@class='v-slide-group__content v-tabs-bar__content']//i[@class='v-icon notranslate v-icon--left mdi mdi-sitemap theme--light']"
    doc_panel_close_button = "//div[@class='v-card__title']//span[@class='headline']/..//i[@class='v-icon notranslate mdi mdi-close theme--light']"
    # document name panel - end
    # document revision click - start
    doc_rev_click = "//a[contains(@href, '#') and text() = '02-NAP-MEP-MDL-000003']/../..//span[@style='cursor: pointer; color: blue;']"
    dev_rev_panel_close = "//button[@class='v-btn v-btn--text theme--light v-size--default secondary--text']//span[@class='v-btn__content' and text()='Close']"

    # document revision click - end

    def __init__(self, driver):
        self.lp = LoginPage(driver)
        self.driver = driver

    # # logger = LogGen.loggen()

    def reg_rows_filter(self):

        #  Method to select the no of documents to be displayed within the reg doc tab
        try:
            WebDriverWait(self.driver, 10).until(expected_conditions.presence_of_element_located(
                (By.XPATH, self.filter_xpath)))
            # Wait for the filter button to be clickable
            WebDriverWait(self.driver, 10).until(expected_conditions.element_to_be_clickable(
                (By.XPATH, self.reg_rows_per_page_filter))).click()
            # logic to move the scroll bar to look for "All" filter
            doc_element = self.driver.find_element(By.XPATH, self.reg_rows_per_page_all)
            ActionChains(self.driver).move_to_element(doc_element).perform()

            WebDriverWait(self.driver, 10).until(expected_conditions.element_to_be_clickable(
                (By.XPATH, self.reg_rows_per_page_all))).click()

        except Exception as e:
            raise e

    def check_doc_properties(self, docTitle):
        # Method to select a document from reg doc tab and create WF
        try:
            # doc_link_xpath = "//a[contains(@href, '#') and text() = '" + docTitle + "']"
            # checkbox_xpath = "//a[contains(@href, '#') and text() = '" + docTitle + "']/../..//div[@class='v-input--selection-controls__ripple']"

            # WebDriverWait(self.driver, 10).until(expected_conditions.presence_of_element_located(
            #     (By.XPATH, self.all_rows_selected)))

            # Scroll the element into view
            doc_element = WebDriverWait(self.driver, 10).until(expected_conditions.element_to_be_clickable(
                (By.XPATH, "//a[contains(@href, '#') and text() = '" + docTitle + "']")))
            ActionChains(self.driver).move_to_element(doc_element).perform()

            WebDriverWait(self.driver, 10).until(expected_conditions.element_to_be_clickable(
                (By.XPATH, "//a[contains(@href, '#') and text() = '" + docTitle + "']"))).click()
            time.sleep(2)
            WebDriverWait(self.driver, 5).until(expected_conditions.element_to_be_clickable(
                (By.XPATH, self.doc_download_panel))).click()
            time.sleep(2)
            WebDriverWait(self.driver, 5).until(expected_conditions.element_to_be_clickable(
                (By.XPATH, self.doc_transmittal_panel))).click()
            time.sleep(2)
            WebDriverWait(self.driver, 5).until(expected_conditions.element_to_be_clickable(
                (By.XPATH, self.doc_wf_panel))).click()
            time.sleep(2)
            WebDriverWait(self.driver, 5).until(expected_conditions.element_to_be_clickable(
                (By.XPATH, self.doc_panel_close_button))).click()

        except Exception as e:
            raise e

    def check_doc_revision(self, docTitle):
        # Method to select a document from reg doc tab and create WF
        try:
            # doc_link_xpath = "//a[contains(@href, '#') and text() = '" + docTitle + "']"
            # revision_xpath = "//a[contains(@href, '#') and text() = '" + docTitle + "']/../..//span[@style='cursor: pointer; color: blue;']"

            # WebDriverWait(self.driver, 10).until(expected_conditions.presence_of_element_located(
            #     (By.XPATH, self.all_rows_selected)))

            # Scroll the element into view
            doc_element = self.driver.find_element(By.XPATH,
                                                   "//a[contains(@href, '#') and text() = '" + docTitle + "']")
            ActionChains(self.driver).move_to_element(doc_element).perform()

            WebDriverWait(self.driver, 10).until(expected_conditions.element_to_be_clickable((By.XPATH,
                                                                                              "//a[contains(@href, '#') and text() = '" + docTitle + "']/../..//span[@style='cursor: pointer; color: blue;']"))).click()
            time.sleep(2)

            WebDriverWait(self.driver, 5).until(expected_conditions.element_to_be_clickable(
                (By.XPATH, self.dev_rev_panel_close))).click()

        except Exception as e:
            raise e


class regDocument_properties:
    reg_rows_per_page_filter = "//div[@name='mass_upload']//div[@class='v-select__selections']//div[@class='v-select__selection v-select__selection--comma' and text()='50']"
    reg_rows_per_page_all = "//div[@class='v-list-item__content']//div[@class='v-list-item__title' and text()='All']"
    filter_xpath = "(//div[@class='v-input--selection-controls__ripple'])[3]"
    restore_xpath = "//div[contains(@class, 'moduleWrapper clearfix ifwe-tabview')]//div[@class='wp-tabview-panel']//div[@class='widget-dd-menu dropdown-menu dropdown-menu-root dropdown dropdown-root']//child::span[@class='maximize-icon fonticon fonticon-resize-small']"
    all_rows_selected = "//div[@class='v-input__slot' and @aria-owns='list-119']//div[@class='v-select__slot']//div[@class='v-select__selection v-select__selection--comma' and text()='All']"
    # document name panel - start
    doc_download_panel = "//div[@class='v-slide-group__content v-tabs-bar__content']//i[@class='v-icon notranslate v-icon--left mdi mdi-account-clock theme--light']"
    doc_transmittal_panel = "//div[@class='v-slide-group__content v-tabs-bar__content']//i[@class='v-icon notranslate v-icon--left mdi mdi-axis-arrow-lock theme--light']"
    doc_wf_panel = "//div[@class='v-slide-group__content v-tabs-bar__content']//i[@class='v-icon notranslate v-icon--left mdi mdi-sitemap theme--light']"
    doc_panel_close_button = "//div[@class='v-card__title']//span[@class='headline']/..//i[@class='v-icon notranslate mdi mdi-close theme--light']"
    # document name panel - end
    # document revision click - start
    doc_rev_click = "//a[contains(@href, '#') and text() = '02-NAP-MEP-MDL-000003']/../..//span[@style='cursor: pointer; color: blue;']"
    dev_rev_panel_close = "//button[@class='v-btn v-btn--text theme--light v-size--default secondary--text']//span[@class='v-btn__content' and text()='Close']"
    iframeNMDocumentCntrl_xpath = "//div[@id='7zAw5w30gCwH0QgEjm01']//iframe[@id='frame-9sMywGd0g06qZr2-U046']"
    get_tra_success_msg = "//div[contains(@class, 'alert-success') and contains(text(), 'Transmittal Created Successfully')]"
    get_tra_no_within_table = "//thead[@class='v-data-table-header']//th[@class='text-start sortable']/../../..//td[@class='text-start' and contains(text(), '-')]"

    # document revision click - end

    def __init__(self, driver):
        self.lp = LoginPage(driver)
        self.driver = driver

    # logger = LogGen.loggen()

    def reg_rows_filter(self):

        #  Method to select the no of documents to be displayed within the reg doc tab
        try:
            self.driver.switch_to.default_content()
            frame1 = WebDriverWait(self.driver, 100).until(expected_conditions.presence_of_element_located(
                (By.XPATH, self.iframeNMDocumentCntrl_xpath)))
            self.driver.switch_to.frame(frame1)

            WebDriverWait(self.driver, 10).until(expected_conditions.presence_of_element_located(
                (By.XPATH, self.filter_xpath)))
            # Wait for the filter button to be clickable
            WebDriverWait(self.driver, 20).until(EC.element_to_be_clickable(
                (By.XPATH, self.reg_rows_per_page_filter))).click()
            # logic to move the scroll bar to look for "All" filter
            doc_element = self.driver.find_element(By.XPATH, self.reg_rows_per_page_all)
            ActionChains(self.driver).move_to_element(doc_element).perform()

            WebDriverWait(self.driver, 10).until(EC.element_to_be_clickable(
                (By.XPATH, self.reg_rows_per_page_all))).click()

        except Exception as e:
            raise e

    # below method currently used to get the TRA number when a TRA is created
    def check_doc_properties(self, docTitle):
        # Method to select a document from reg doc tab and create WF
        try:
            doc_link_xpath = "//a[contains(@href, '#') and text() = '" + docTitle + "']"

            # WebDriverWait(self.driver, 10).until(expected_conditions.presence_of_element_located(
            #     (By.XPATH, self.all_rows_selected)))

            # Scroll the element into view
            doc_element = self.driver.find_element(By.XPATH, doc_link_xpath)
            ActionChains(self.driver).move_to_element(doc_element).perform()

            WebDriverWait(self.driver, 10).until(EC.element_to_be_clickable((By.XPATH, doc_link_xpath))).click()
            time.sleep(2)
            # WebDriverWait(self.driver, 5).until(expected_conditions.element_to_be_clickable(
            #     (By.XPATH, self.doc_download_panel))).click()
            time.sleep(2)
            WebDriverWait(self.driver, 5).until(expected_conditions.element_to_be_clickable(
                (By.XPATH, self.doc_transmittal_panel))).click()
            time.sleep(2)

            tra_no_from_table = WebDriverWait(self.driver, 5).until(expected_conditions.visibility_of_element_located(
                (By.XPATH, self.get_tra_no_within_table)))
            tra_no_text = tra_no_from_table.text
            print("tra no is ", tra_no_text)

            file_path = ".//TestData//DataManager.xlsx"

            # Load the existing workbook
            workbook = openpyxl.load_workbook(file_path)

            # Select the specific sheet by name
            sheet = workbook["transmittal"]

            # Assuming you have extracted the text value
            # extracted_text = "PMC-RFI-000006"

            # Write the extracted text to cell D3
            sheet.cell(row=3, column=4, value=tra_no_text)

            # Save the changes to the workbook
            workbook.save(file_path)

            # WebDriverWait(self.driver, 5).until(expected_conditions.element_to_be_clickable(
            #     (By.XPATH, self.doc_wf_panel))).click()
            time.sleep(2)
            WebDriverWait(self.driver, 5).until(expected_conditions.element_to_be_clickable(
                (By.XPATH, self.doc_panel_close_button))).click()

        except Exception as e:
            raise e

    def check_doc_revision(self, docTitle):
        # Method to select a document from reg doc tab and create WF
        try:
            doc_link_xpath = "//a[contains(@href, '#') and text() = '" + docTitle + "']"
            revision_xpath = "//a[contains(@href, '#') and text() = '" + docTitle + "']/../..//span[@style='cursor: pointer; color: blue;']"

            # WebDriverWait(self.driver, 10).until(expected_conditions.presence_of_element_located(
            #     (By.XPATH, self.all_rows_selected)))

            # Scroll the element into view
            doc_element = self.driver.find_element(By.XPATH, doc_link_xpath)
            ActionChains(self.driver).move_to_element(doc_element).perform()

            WebDriverWait(self.driver, 10).until(EC.element_to_be_clickable((By.XPATH, revision_xpath))).click()
            time.sleep(2)

            WebDriverWait(self.driver, 5).until(expected_conditions.element_to_be_clickable(
                (By.XPATH, self.dev_rev_panel_close))).click()

        except Exception as e:
            raise e

    # this method is to handle the the preference menu "already opened"

    # def click_menu_dropdown(self, widget_name):
    #     try:
    #         WebDriverWait(self.driver, 10).until(expected_conditions.presence_of_element_located(
    #             (By.XPATH, self.restore_xpath)))
    #         WebDriverWait(self.driver, 10).until(expected_conditions.element_to_be_clickable(
    #             (By.XPATH,
    #              "//div[contains(@class, 'moduleWrapper clearfix ifwe-tabview')]//div[@class='wp-tabview-panel']//child::div[contains(@class,'moduleHeader__title') and text()='" + widget_name + "']//following-sibling::div//child::span[@class='widget-menu-icon ifwe-action-icon fonticon fonticon-down-open clickable active']"))).click()
    #         # self.logger.info("** widget is already maximized **")
    #         time.sleep(5)
    #     except Exception as e:
    #         self.lp.maximize_widget()
    #         # self.logger.info("** widget is maximized **")
    #         time.sleep(5)
    #         raise e


class Create_Transmittal:
    iframeNMDocumentCntrl_xpath = "//div[@id='7zAw5w30gCwH0QgEjm01']//iframe[@id='frame-9sMywGd0g06qZr2-U046']"
    mail_management_iframe_xpath = "//div[@class='moduleWrapper']//iframe[@id='frame-9wOjNo70g06qNYY8Lm01']"
    plus_menu_button = "//i[@class='v-icon notranslate mdi mdi-plus theme--dark']"
    # create_workflow_button = "//button[@class='v-btn v-btn--is-elevated v-btn--fab v-btn--has-bg v-btn--round theme--dark v-size--small yellow']//div[contains(@style, 'Create%20Workflow')]"
    widget_name = 'Document Register'
    reg_rows_per_page_filter = "//div[@class='v-input__slot' and @aria-owns='list-119']//div[@class='v-select__slot']//div[@class='v-select__selection v-select__selection--comma' and text()='50']"
    # reg_rows_per_page_filter = "//div[@class='v-data-footer__icons-after']//button[@class='v-btn v-btn--icon v-btn--round v-btn--text theme--light v-size--default']"
    reg_rows_per_page_all = "//div[@class='v-list-item__content']//div[@class='v-list-item__title' and text()='All']"
    all_rows_selected = "//div[@class='v-input__slot' and @aria-owns='list-119']//div[@class='v-select__slot']//div[@class='v-select__selection v-select__selection--comma' and text()='All']"

    # transmittal START -----------------
    create_transmittal_xpath = "//div[@style='transition-delay: 0.1s;']//div[@class='v-responsive__content']"
    # tra_type_gcd = "//select[@name='IMP_MailType']//option[@value='GCD']"
    # tra_type_trans = "//select[@name='IMP_MailType']//option[@value='TRA']"
    # tra_type_rfi = "//select[@name='IMP_MailType']//option[@value='RFI']"
    tra_to_list = "//div[@linked='IMP_ToList']"
    # tra_to_person_selection = "//span[@class='searchItemSpan search_item_in_apps' and text()='Himanshu Sharma']/../../../../..//div[@class='wux-controls-responsivetileview-checkbox']"
    # tra_to_person_selection = "//div[@class='wux-controls-abstract wux-controls-responsivetileview wux-ui-state-activated wux-ui-state-prehighlighted']//div[@class='wux-controls-highlightborder wux-ui-state-prehighlighted']"
    tra_person_ok_button = "//button[@id='id_in_app_ok']"
    tra_cc_list = "//div[@linked='IMP_CCList']"
    # tra_cc_person_selection = "//span[@class='searchItemSpan search_item_in_apps' and text()='Jayesh Panat']/../../../../..//div[@class='wux-controls-responsivetileview-checkbox']"
    tra_reason = "//select[@name='IMP_ReasonForIssue']"
    # tra_response = "//select[@name='IMP_MailResponseRequired']//option[@value='Yes']"
    tra_DatePicker = "//input[@id='calendar_IMP_MailResponseRequiredDate']"
    tra_subject = "//input[@name='IMP_Subject']"
    tra_category = "//select[@name='IMP_TRACategory']"
    tra_contract = "//select[@name='IMP_ContractCode']"
    tra_workorder = "//select[@name='IMP_Workorder']"
    tra_message = "//body[@aria-label='Editor, IMP_Message']//p"
    tra_send_button = "//div[@style='text-align: right; height: 70px; padding: 17.5px; background-color: rgb(241, 241, 241); border-bottom-left-radius: 6px; border-bottom-right-radius: 6px; border-top: 1px solid rgb(229, 229, 229);']//button[@class='btn-primary btn hd-form-submit-hover hd-form-submit-focus' and text()='Send']"
    tra_success_msg = "//div[contains(@class, 'alert-success') and contains(text(), 'Transmittal Created Successfully')]"
    # --- for TRA
    tra_response_value = "//select[@name='IMP_ReasonForIssue']//option[@value='IFD']"
    # tra_category_value = "//select[@name='IMP_TRACategory']//option[@value='MeetingMinutes']"
    tra_letter_category = "//select[@name='IMP_TRACategory']//option[@value='Letter']"
    # tra_contract_value = "//select[@name='IMP_ContractCode']//option[@value='0000100113']"
    # tra_wo_value = "//select[@name='IMP_Workorder']//option[@value='025']"
    tra_discipline_field_click = "//div[@style='font-size: 14px; top: 3px; position: relative; padding: 4px; text-overflow: ellipsis; overflow: hidden; height: 30px; color: rgb(119, 121, 124); font-weight: 700;' and text()='Discipline']/../..//div[@class='v-input__control']//div[@class='v-select__selections']"
    tra_ltr_ini_field_check = "//div[@style='font-size: 14px; top: 3px; position: relative; padding: 4px; text-overflow: ellipsis; overflow: hidden; height: 30px; color: rgb(119, 121, 124); font-weight: 700;' and text()='Letter Initiator']/../..//div[@class='v-input__control']//div[@class='v-select__selections']"
    tra_ltr_adr_field_check = "//div[@style='font-size: 14px; top: 3px; position: relative; padding: 4px; text-overflow: ellipsis; overflow: hidden; height: 30px; color: rgb(119, 121, 124); font-weight: 700;' and text()='Letter Addressee']/../..//div[@class='v-input__control']//div[@class='v-select__selections']"
    message_iframe_xpath = "//iframe[@class='cke_wysiwyg_frame cke_reset']"
    locate_mesg_field = "//section[@style='width: 98%; padding-top: 3px; padding-bottom: 3px;']//h5[@style='padding-left: 10px;' and text()='Message']"
    locate_category_field = "//div[@style='font-size: 14px; top: 3px; position: relative; padding: 4px; text-overflow: ellipsis; overflow: hidden; height: 30px; color: rgb(119, 121, 124); font-weight: 700;' and text()='Transmittal Category']"
    locate_contract_field = "//div[@style='font-size: 14px; top: 3px; position: relative; padding: 4px; text-overflow: ellipsis; overflow: hidden; height: 30px; color: rgb(119, 121, 124); font-weight: 700;' and text()='Contract']"
    locate_wo_field = "//div[@style='font-size: 14px; top: 3px; position: relative; padding: 4px; text-overflow: ellipsis; overflow: hidden; height: 30px; color: rgb(119, 121, 124); font-weight: 700;' and text()='Work Order/Service Order']"
    locate_reason_field = "//div[@style='font-size: 14px; top: 3px; position: relative; padding: 4px; text-overflow: ellipsis; overflow: hidden; height: 30px; color: rgb(119, 121, 124); font-weight: 700;' and text()='Reason For Issue']"
    locate_discipline_field = "//div[@style='font-size: 14px; top: 3px; position: relative; padding: 4px; text-overflow: ellipsis; overflow: hidden; height: 30px; color: rgb(119, 121, 124); font-weight: 700;' and text()='Discipline']"
    locate_ltr_ini_field = "//div[@style='font-size: 14px; top: 3px; position: relative; padding: 4px; text-overflow: ellipsis; overflow: hidden; height: 30px; color: rgb(119, 121, 124); font-weight: 700;' and text()='Letter Initiator']"
    locate_ltr_adr_field = "//div[@style='font-size: 14px; top: 3px; position: relative; padding: 4px; text-overflow: ellipsis; overflow: hidden; height: 30px; color: rgb(119, 121, 124); font-weight: 700;' and text()='Letter Addressee']"
    disc_display_list = "//div[@class='v-list v-select-list v-sheet theme--light v-list--dense theme--light' and @role='listbox']"
    ltr_initiator_list = "//div[@class='v-menu__content theme--light menuable__content__active v-autocomplete__content']//div[@class='v-list v-select-list v-sheet theme--light v-list--dense theme--light' and @role='listbox']"
    ltr_addressee_list = "//div[@class='v-menu__content theme--light menuable__content__active v-autocomplete__content']//div[@class='v-list v-select-list v-sheet theme--light v-list--dense theme--light' and @role='listbox']"

    # transmittal END ------------------
    # RFI START----------------------------------
    # rfi_assetcode_value = "//select[@name='IMP_Reg02_AssetCodeLevel6']//option[@value = '012110']"
    locate_asset_code_field = "//div[@style='font-size: 14px; top: 3px; position: relative; padding: 4px; text-overflow: ellipsis; overflow: hidden; height: 30px; color: rgb(119, 121, 124); font-weight: 700;' and text()='Asset Code']"
    locate_response_code_field = "//div[@style='font-size: 14px; top: 3px; position: relative; padding: 4px; text-overflow: ellipsis; overflow: hidden; height: 30px; color: rgb(119, 121, 124); font-weight: 700;' and text()='Response Required']"
    # rfi_stage_value = "//select[@name='IMP_StageCode']//option[@value='02A']"
    locate_stage_field = "//div[@style='font-size: 14px; top: 3px; position: relative; padding: 4px; text-overflow: ellipsis; overflow: hidden; height: 30px; color: rgb(119, 121, 124); font-weight: 700;' and text()='Stage']"
    # rfi_design_value = "//select[@name='IMP_RFIDesignRelated']//option[@value='Yes']"
    locate_design_field = "//div[@style='font-size: 14px; top: 3px; position: relative; padding: 4px; text-overflow: ellipsis; overflow: hidden; height: 30px; color: rgb(119, 121, 124); font-weight: 700;' and text()='Design Related']"
    rfi_info_value = "//textarea[@name='IMP_RFIInformationRequested']"
    locate_info_field = "//div[@style='font-size: 14px; top: 3px; position: relative; padding: 4px; text-overflow: ellipsis; overflow: hidden; height: 30px; color: rgb(119, 121, 124); font-weight: 700;' and text()='Information Requested']"

    # constructor to initialise driver
    def __init__(self, driver):
        self.lp = LoginPage(driver)
        self.driver = driver
        self.dp = regDocument_properties(self.driver)

    # logger = LogGen.loggen()

    def create_transmittal_button(self):
        try:

            WebDriverWait(self.driver, 10).until(expected_conditions.element_to_be_clickable(
                (By.XPATH, self.plus_menu_button))).click()
            WebDriverWait(self.driver, 10).until(expected_conditions.element_to_be_clickable(
                (By.XPATH, self.create_transmittal_xpath))).click()
        except Exception as e:
            raise e

    # Transmittal Type-General Correspondence -- START
    def traType_gcd(self, type):
        try:
            try:
                # WebDriverWait(self.driver, 5).until(expected_conditions.element_to_be_clickable(
                #     (By.XPATH, self.tra_enterType))).send_keys(traType)
                tra_selection = WebDriverWait(self.driver, 10).until(expected_conditions.presence_of_element_located(
                    (By.XPATH, "//select[@name='IMP_MailType']//option[@value='" + type + "']")))
                tra_selection.click()
            except:
                pass
        except Exception as e:
            raise e

    def traType_rfi(self, value):
        try:
            try:
                # WebDriverWait(self.driver, 5).until(expected_conditions.element_to_be_clickable(
                #     (By.XPATH, self.tra_enterType))).send_keys(traType)
                # tra_selection = WebDriverWait(self.driver, 10).until(expected_conditions.presence_of_element_located(
                #     (By.XPATH, self.tra_type_rfi)))
                tra_selection = WebDriverWait(self.driver, 10).until(expected_conditions.presence_of_element_located(
                    (By.XPATH, "//select[@name='IMP_MailType']//option[@value='" + value + "']")))
                tra_selection.click()
                # dropdown_element = Select(tra_selection)
                # dropdown_element.select_by_visible_text(traType)
            except:
                pass
        except Exception as e:
            raise e

    def tra_toUser(self, user):
        try:

            WebDriverWait(self.driver, 10).until(expected_conditions.element_to_be_clickable(
                (By.XPATH, self.tra_to_list))).click()
            time.sleep(15)
            self.driver.switch_to.default_content()
            WebDriverWait(self.driver, 20).until(expected_conditions.element_to_be_clickable(
                (By.XPATH,
                 "//div[@class='wux-controls-abstract wux-controls-responsivetileview']//div[@class='wux-controls-responsivetileview-maincontent']//div[@class='wux-button-label-ellipsis-wrap wux-ui-is-rendered wux-controls-responsivetileview-description']//span[@class='searchItemSpan' and text()='" + user + "']"))).click()
            # self.logger.info("**** Person selected ")
            WebDriverWait(self.driver, 5).until(expected_conditions.element_to_be_clickable(
                (By.XPATH, self.tra_person_ok_button))).click()
            time.sleep(5)
        except Exception as e:
            raise e

    def tra_ccUser(self, user):
        try:
            self.driver.switch_to.default_content()
            frame1 = WebDriverWait(self.driver, 10).until(expected_conditions.presence_of_element_located(
                (By.XPATH, self.iframeNMDocumentCntrl_xpath)))
            self.driver.switch_to.frame(frame1)

            WebDriverWait(self.driver, 10).until(expected_conditions.element_to_be_clickable(
                (By.XPATH, self.tra_cc_list))).click()
            time.sleep(15)
            self.driver.switch_to.default_content()
            WebDriverWait(self.driver, 20).until(expected_conditions.element_to_be_clickable(
                (By.XPATH,
                 "//div[@class='wux-controls-abstract wux-controls-responsivetileview']//div[@class='wux-controls-responsivetileview-maincontent']//div[@class='wux-button-label-ellipsis-wrap wux-ui-is-rendered wux-controls-responsivetileview-description']//span[@class='searchItemSpan' and text()='" + user + "']"))).click()
            # self.logger.info("**** Template selected ")
            WebDriverWait(self.driver, 5).until(expected_conditions.element_to_be_clickable(
                (By.XPATH, self.tra_person_ok_button))).click()
            time.sleep(5)
        except Exception as e:
            raise e

    def traResponseRequired(self, value):
        try:

            # self.driver.switch_to.default_content()
            # use code when creating the transmittal - START
            # frame1 = WebDriverWait(self.driver, 10).until(expected_conditions.presence_of_element_located(
            #     (By.XPATH, self.iframeNMDocumentCntrl_xpath)))
            # self.driver.switch_to.frame(frame1)
            # use code when creating the transmittal - END

            # frame_mail = WebDriverWait(self.driver, 10).until(expected_conditions.presence_of_element_located(
            #     (By.XPATH, self.mail_management_iframe_xpath)))
            # self.driver.switch_to.frame(frame_mail)

            response_field = WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.XPATH, self.locate_response_code_field)))

            self.driver.execute_script("arguments[0].scrollIntoView(true);", response_field)

            tra_responserqd = WebDriverWait(self.driver, 5).until(expected_conditions.element_to_be_clickable(
                (By.XPATH, "//select[@name='IMP_MailResponseRequired']//option[@value='" + value + "']")))
            tra_responserqd.click()

        except Exception as e:
            raise e

    def traResponseDatePicker(self):
        try:
            date_picker = WebDriverWait(self.driver, 5).until(expected_conditions.element_to_be_clickable(
                (By.XPATH, self.tra_DatePicker)))
            date_picker.click()
            time.sleep(1)

            date_picker.send_keys(Keys.ENTER)

        except Exception as e:
            raise e

    def tra_Subject(self, value):
        #  Method to input WF details in WF creation window
        try:
            self.driver.switch_to.default_content()
            frame1 = WebDriverWait(self.driver, 10).until(expected_conditions.presence_of_element_located(
                (By.XPATH, self.iframeNMDocumentCntrl_xpath)))
            self.driver.switch_to.frame(frame1)
            WebDriverWait(self.driver, 10).until(expected_conditions.presence_of_element_located(
                (By.XPATH, self.tra_subject))).send_keys(value)
            time.sleep(5)
        except Exception as e:
            raise e

    def tra_Message(self, msg):

        try:
            self.driver.switch_to.default_content()
            frame1 = WebDriverWait(self.driver, 10).until(expected_conditions.presence_of_element_located(
                (By.XPATH, self.iframeNMDocumentCntrl_xpath)))
            self.driver.switch_to.frame(frame1)
            # self.logger.info("in frame 1")
            message_field = WebDriverWait(self.driver, 2000).until(
                EC.presence_of_element_located((By.XPATH, self.locate_mesg_field)))
            self.driver.execute_script("arguments[0].scrollIntoView(true);", message_field)
            frame2 = WebDriverWait(self.driver, 10).until(expected_conditions.presence_of_element_located(
                (By.XPATH, "//div[@class='cke_contents cke_reset']//iframe[@class='cke_wysiwyg_frame cke_reset']")))
            self.driver.switch_to.frame(frame2)
            # self.logger.info("in frame 2")
            message = WebDriverWait(self.driver, 10).until(expected_conditions.presence_of_element_located(
                (By.XPATH,
                 "//html[@dir='ltr']//body[@class='cke_editable cke_editable_themed cke_contents_ltr cke_show_borders']")))
            # self.logger.info("in Message field")
            message.send_keys(msg)
            time.sleep(10)
        except Exception as e:
            raise e

    # Transmittal Type-General Correspondence -- END

    # Transmittal Type-TRANSMITTAL -- START
    def traType_transmittal(self, type):
        try:
            try:
                # WebDriverWait(self.driver, 5).until(expected_conditions.element_to_be_clickable(
                #     (By.XPATH, self.tra_enterType))).send_keys(traType)
                tra_selection = WebDriverWait(self.driver, 10).until(expected_conditions.presence_of_element_located(
                    (By.XPATH, "//select[@name='IMP_MailType']//option[@value='" + type + "']")))
                tra_selection.click()
                # dropdown_element = Select(tra_selection)
                # dropdown_element.select_by_visible_text(traType)
            except:
                pass
        except Exception as e:
            raise e

    # Transmittal Type-TRANSMITTAL -- START

    def traReason(self, reason):
        try:
            # always change the frames code when switching between CREATE TRA or REPLY/FORWARD TRA

            # self.driver.switch_to.default_content()
            # frame1 = WebDriverWait(self.driver, 10).until(expected_conditions.presence_of_element_located(
            #     (By.XPATH, self.iframeNMDocumentCntrl_xpath)))
            # self.driver.switch_to.frame(frame1)
            self.driver.switch_to.default_content()
            mail_frame = WebDriverWait(self.driver, 10).until(expected_conditions.presence_of_element_located(
                (By.XPATH, self.mail_management_iframe_xpath)))
            self.driver.switch_to.frame(mail_frame)
            reason_field = WebDriverWait(self.driver, 20).until(
                EC.presence_of_element_located((By.XPATH, self.locate_reason_field)))
            self.driver.execute_script("arguments[0].scrollIntoView(true);", reason_field)
            reason_select = WebDriverWait(self.driver, 5).until(expected_conditions.element_to_be_clickable(
                (By.XPATH, "//select[@name='IMP_ReasonForIssue']//option[@value='" + reason + "']")))
            reason_select.click()
        except Exception as e:
            raise e

    def tra_category_other(self, category_val):

        try:
            category_field = WebDriverWait(self.driver, 20).until(
                EC.presence_of_element_located((By.XPATH, self.locate_category_field)))
            self.driver.execute_script("arguments[0].scrollIntoView(true);", category_field)

            category_other_select = WebDriverWait(self.driver, 10).until(
                expected_conditions.presence_of_element_located(
                    (By.XPATH, "//select[@name='IMP_TRACategory']//option[@value='" + category_val + "']")))
            category_other_select.click()

        except Exception as e:
            raise e

    def tra_category_ltr(self, category_val):

        try:
            category_field = WebDriverWait(self.driver, 20).until(
                EC.presence_of_element_located((By.XPATH, self.locate_category_field)))
            self.driver.execute_script("arguments[0].scrollIntoView(true);", category_field)

            category_letter_select = WebDriverWait(self.driver, 10).until(
                expected_conditions.presence_of_element_located(
                    (By.XPATH, "//select[@name='IMP_TRACategory']//option[@value='" + category_val + "']")))
            category_letter_select.click()

        except Exception as e:
            raise e

    def ltr_initiator(self, ltr_ini):
        try:
            ltr_ini_field = WebDriverWait(self.driver, 20).until(
                EC.presence_of_element_located((By.XPATH, self.locate_ltr_ini_field)))
            self.driver.execute_script("arguments[0].scrollIntoView(true);", ltr_ini_field)

            ltr_ini_select = WebDriverWait(self.driver, 10).until(expected_conditions.presence_of_element_located(
                (By.XPATH, self.tra_ltr_ini_field_check)))
            ltr_ini_select.click()

            # Locate the list of initiators
            WebDriverWait(self.driver, 30).until(expected_conditions.visibility_of_element_located(
                (By.XPATH, self.ltr_initiator_list)))

            # self.logger.info("discipline value is %s", discipline_value)
            ltr_ini_value = str(ltr_ini).strip()  # Ensure discipline_value is converted to a string
            ltr_ini_link_xpath = f"//div[@class='v-list-item__title' and text()='{ltr_ini_value}']"

            initiator_link_element = WebDriverWait(self.driver, 10).until(
                EC.visibility_of_element_located((By.XPATH, ltr_ini_link_xpath)))
            print("initiator luink, ", initiator_link_element)
            # Scroll the element into view
            ActionChains(self.driver).move_to_element(initiator_link_element).perform()

            checkbox_xpath = "//div[@class='v-list-item__title' and text()='" + ltr_ini_value + "']/../..//div[@class='v-simple-checkbox']"
            WebDriverWait(self.driver, 10).until(EC.element_to_be_clickable((By.XPATH, checkbox_xpath))).click()

            time.sleep(1)
            ini_click_event_xpath = "//div[@style='font-size: 14px; top: 3px; position: relative; padding: 4px; text-overflow: ellipsis; overflow: hidden; height: 30px; color: rgb(119, 121, 124); font-weight: 700;' and contains(text(), 'Letter Initiator')]"
            ini_close_selection = self.driver.find_element(By.XPATH, ini_click_event_xpath)

            actions = ActionChains(self.driver)
            # Perform left-click on the element
            actions.click(ini_close_selection).perform()
        except Exception as e:
            raise e

    # def ltr_addressee(self, ltr_adr):
    #     try:
    #         ltr_adr_field = WebDriverWait(self.driver, 20).until(
    #             EC.presence_of_element_located((By.XPATH, self.locate_ltr_adr_field)))
    #         self.driver.execute_script("arguments[0].scrollIntoView(true);", ltr_adr_field)
    #
    #         ltr_adr_select = WebDriverWait(self.driver, 10).until(
    #             EC.presence_of_element_located((By.XPATH, self.tra_ltr_adr_field_check)))
    #         ltr_adr_select.click()
    #
    #         # Wait for the list of addressees to be visible
    #         WebDriverWait(self.driver, 30).until(
    #             EC.visibility_of_element_located((By.XPATH, self.ltr_addressee_list)))
    #
    #         ltr_add_value = str(ltr_adr).strip()  # Convert to string
    #         ltr_add_link_xpath = f"//div[@class='v-list-item__title' and text()='{ltr_add_value}']"
    #
    #         # Scroll to the element before locating
    #         self.driver.execute_script("arguments[0].scrollIntoView(true);", ltr_adr_field)
    #
    #         # Wait for the element to be clickable
    #         addressee_link_element = WebDriverWait(self.driver, 20).until(
    #             EC.element_to_be_clickable((By.XPATH, ltr_add_link_xpath)))
    #
    #         # Perform a click on the addressee link
    #         addressee_link_element.click()
    #
    #         time.sleep(1)
    #
    #         # Locate and click the close selection button
    #         adr_click_event_xpath = "//div[contains(text(), 'Letter Addressee') and @style='font-size: 14px; top: 3px; position: relative; padding: 4px; text-overflow: ellipsis; overflow: hidden; height: 30px; color: rgb(119, 121, 124); font-weight: 700;']"
    #         adr_close_selection = WebDriverWait(self.driver, 10).until(
    #             EC.element_to_be_clickable((By.XPATH, adr_click_event_xpath)))
    #
    #         actions = ActionChains(self.driver)
    #         actions.click(adr_close_selection).perform()
    #     except Exception as e:
    #         raise e

    def ltr_addressee(self, ltr_adr):
        try:
            ltr_adr_field = WebDriverWait(self.driver, 20).until(
                EC.visibility_of_element_located((By.XPATH, self.locate_ltr_adr_field)))
            self.driver.execute_script("arguments[0].scrollIntoView(true);", ltr_adr_field)

            ltr_adr_select = WebDriverWait(self.driver, 10).until(
                EC.visibility_of_element_located((By.XPATH, self.tra_ltr_adr_field_check)))
            ltr_adr_select.click()

            for value in ltr_adr:
                ltr_adr_link_xpath = f"//div[@class='v-list-item__content']//div[@class='v-list-item__title' and text()='{value}']"
                checkbox_xpath = f"{ltr_adr_link_xpath}/../..//div[@class='v-input--selection-controls__input']"
                print("ltr add xpath ", ltr_adr_link_xpath)
                # Scroll the element into view
                doc_element = self.driver.find_element(By.XPATH, ltr_adr_link_xpath)
                ActionChains(self.driver).move_to_element(doc_element).perform()

                WebDriverWait(self.driver, 10).until(EC.element_to_be_clickable((By.XPATH, checkbox_xpath))).click()

            time.sleep(1)
            adr_click_event_xpath = "//div[contains(text(), 'Letter Addressee')]"
            adr_close_selection = self.driver.find_element(By.XPATH, adr_click_event_xpath)

            actions = ActionChains(self.driver)
            # Perform left-click on the element
            actions.click(adr_close_selection).perform()
        except Exception as e:
            print(f"Error occurred: {e}")
            raise e

    def tra_contract(self, value):

        try:
            message_field = WebDriverWait(self.driver, 20).until(
                EC.presence_of_element_located((By.XPATH, self.locate_contract_field)))
            self.driver.execute_script("arguments[0].scrollIntoView(true);", message_field)

            # contract_select = WebDriverWait(self.driver, 10).until(expected_conditions.presence_of_element_located(
            #     (By.XPATH, self.tra_contract_value)))
            contract_select = WebDriverWait(self.driver, 10).until(expected_conditions.presence_of_element_located(
                (By.XPATH, "//select[@name='IMP_ContractCode']//option[@value='" + value + "']")))
            contract_select.click()

        except Exception as e:
            raise e

    def tra_wo(self, value):

        try:
            message_field = WebDriverWait(self.driver, 20).until(
                EC.presence_of_element_located((By.XPATH, self.locate_wo_field)))
            self.driver.execute_script("arguments[0].scrollIntoView(true);", message_field)

            # wo_select = WebDriverWait(self.driver, 10).until(expected_conditions.presence_of_element_located(
            #     (By.XPATH, self.tra_wo_value)))
            wo_select = WebDriverWait(self.driver, 10).until(expected_conditions.presence_of_element_located(
                (By.XPATH, "//select[@name='IMP_Workorder']//option[@value='" + value + "']")))
            wo_select.click()

        except Exception as e:
            raise e

    def tra_discipline(self, discipline_value):
        try:
            desc_field = WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.XPATH, self.locate_discipline_field)))
            self.driver.execute_script("arguments[0].scrollIntoView(true);", desc_field)

            discipline_select = WebDriverWait(self.driver, 10).until(
                expected_conditions.element_to_be_clickable((By.XPATH, self.tra_discipline_field_click)))
            discipline_select.click()

            # Locate the list of disciplines
            WebDriverWait(self.driver, 30).until(expected_conditions.visibility_of_element_located(
                (By.XPATH, self.disc_display_list)))

            # self.logger.info("discipline value is %s", discipline_value)
            discipline = str(discipline_value).strip()  # Ensure discipline_value is converted to a string
            discipline_link_xpath = f"//div[@class='v-list-item__title' and text()='{discipline}']"
            discipline_link_element = WebDriverWait(self.driver, 10).until(
                EC.visibility_of_element_located((By.XPATH, discipline_link_xpath)))

            # Scroll the element into view
            ActionChains(self.driver).move_to_element(discipline_link_element).perform()

            # Click the checkbox
            checkbox_xpath = "//div[@class='v-list-item__title' and text()='" + discipline + "']/../..//div[@class='v-simple-checkbox']"
            WebDriverWait(self.driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, checkbox_xpath))).click()

            time.sleep(1)
            # Close the discipline selection
            disc_click_event_xpath = "//div[contains(text(), 'Discipline') and @style='font-size: 14px; top: 3px; position: relative; padding: 4px; text-overflow: ellipsis; overflow: hidden; height: 30px; color: rgb(119, 121, 124); font-weight: 700;']"
            discipline_close_selection = WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.XPATH, disc_click_event_xpath)))

            actions = ActionChains(self.driver)
            actions.click(discipline_close_selection).perform()

        except Exception as e:
            raise e

    # Transmittal Type-TRANSMITTAL -- END

    # Transmittal Type-RFI -- START
    # Transmittal Type-RFI -- END

    def tra_send(self, docTitle):
        try:
            self.driver.switch_to.default_content()
            frame1 = WebDriverWait(self.driver, 10).until(expected_conditions.presence_of_element_located(
                (By.XPATH, self.iframeNMDocumentCntrl_xpath)))
            self.driver.switch_to.frame(frame1)

            send_button = WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.XPATH, self.tra_send_button)))
            send_button.click()

            # Wait for the success message to be displayed
            WebDriverWait(self.driver, 10).until(
                EC.visibility_of_element_located((By.XPATH, self.tra_success_msg)))

            # Execute the check_doc_properties method after the success message is displayed
            self.dp.check_doc_properties(docTitle)
        except Exception as e:
            raise e

    # -------- Transmittal Management END--------------------------------------------

    # ----------------------------------- RFI attributes START-----------------------------
    def rfi_assetCode(self, value):
        try:
            asset_field = WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.XPATH, self.locate_asset_code_field)))
            self.driver.execute_script("arguments[0].scrollIntoView(true);", asset_field)

            # Convert the integer value to a string before concatenating
            value_str = str(value)
            asset_value = WebDriverWait(self.driver, 10).until(expected_conditions.presence_of_element_located(
                (By.XPATH, "//select[@name='IMP_Reg02_AssetCodeLevel6']//option[@value = '" + value_str + "']")))
            asset_value.click()

        except Exception as e:
            print("Error:", e)
            raise e

    def rfi_stage(self, value):
        try:
            stage_field = WebDriverWait(self.driver, 20).until(
                EC.presence_of_element_located((By.XPATH, self.locate_stage_field)))
            self.driver.execute_script("arguments[0].scrollIntoView(true);", stage_field)

            # asset_value = WebDriverWait(self.driver, 5).until(expected_conditions.presence_of_element_located(
            #     (By.XPATH, self.rfi_stage_value)))
            stage_value = WebDriverWait(self.driver, 5).until(expected_conditions.presence_of_element_located(
                (By.XPATH, "//select[@name='IMP_StageCode']//option[@value= '" + value + "']")))
            stage_value.click()

        except Exception as e:
            raise e

    def rfi_design(self, value):
        #  Method to input WF details in WF creation window
        try:
            # WebDriverWait(self.driver, 5).until(expected_conditions.element_to_be_clickable(
            #     (By.XPATH, self.rfi_design_value))).click()
            WebDriverWait(self.driver, 5).until(expected_conditions.element_to_be_clickable(
                (By.XPATH, "//select[@name='IMP_RFIDesignRelated']//option[@value= '" + value + "']"))).click()
            time.sleep(5)
        except Exception as e:
            raise e

    def rfi_info_requested(self, value):
        #  Method to input WF details in WF creation window
        try:
            # WebDriverWait(self.driver, 10).until(expected_conditions.presence_of_element_located(
            #     (By.XPATH, self.rfi_info_value))).send_keys("Test Information")
            WebDriverWait(self.driver, 10).until(expected_conditions.presence_of_element_located(
                (By.XPATH, self.rfi_info_value))).send_keys(value)
            time.sleep(5)
        except Exception as e:
            raise e

    # ----------------------------------- RFI attributes END -----------------------------

    # Verifying the TRANSMITAL & Replying to a transmittal - START ===========================================================


class transmittal_reply:
    mail_management_iframe_xpath = "//div[@class='moduleWrapper']//iframe[@id='frame-9wOjNo70g06qNYY8Lm01']"
    filter_button = "//div[@style='z-index: 999999; height: 35px; margin-bottom: 5px; width: 100%;']//span[@class='fa fa-filter dropbtn']"
    filter_value_outstanding = "//div[@id='myDropdown']//div[@value='Outstanding']"
    filter_value_NA = "//div[@id='myDropdown']//div[@value='NA']"
    filter_value_Sent = "//div[@id='myDropdown']//div[@value='OwnedByMe']"
    filter_value_Inbox = "//div[@id='myDropdown']//div[@value='AssignedToMe']"
    close_filter_window_event = "//div[@style='z-index: 999999; height: 35px; margin-bottom: 5px; width: 100%;']"
    tra_open_button = "//ul[@class='custom-menu']//li[@title='Open']//div[contains(text(), 'Open')]"
    mail_properties_label = "//div[@class='floatingPanel_Header']//span[@class='floatingPanel_title']//span[contains(text(), 'Mail Properties')]"
    reply_button_xpath = "//div[@style='text-align: right; height: 70px; padding: 17.5px; background-color: rgb(241, 241, 241); border-bottom-left-radius: 6px; border-bottom-right-radius: 6px; border-top: 1px solid rgb(229, 229, 229);']//button[@type='button' and text()='Reply']"
    tra_response = "//select[@name='IMP_MailResponseRequired']//option[@value='Yes']"
    locate_mesg_field = "//section[@style='width: 98%; padding-top: 3px; padding-bottom: 3px;']//h5[@style='padding-left: 10px;' and text()='Message']"
    tra_person_ok_button = "//button[@id='id_in_app_ok']"
    tra_cc_list = "//div[@linked='IMP_CCList']"
    tra_to_list = "//div[@linked='IMP_ToList']"
    forward_button_xpath = "//button[@class='btn-primary btn hd-form-submit-hover hd-form-submit-focus' and contains(@style, 'background-color: rgb(54, 142, 196)') and contains(@style, 'border-color: rgb(54, 142, 196)') and text()='Forward']"
    tra_discipline_field_click = "//div[@style='font-size: 14px; top: 3px; position: relative; padding: 4px; text-overflow: ellipsis; overflow: hidden; height: 30px; color: rgb(119, 121, 124); font-weight: 700;' and text()='Discipline']/../..//div[@class='v-input__control']//div[@class='v-select__selections']"
    locate_discipline_field = "//div[@style='font-size: 14px; top: 3px; position: relative; padding: 4px; text-overflow: ellipsis; overflow: hidden; height: 30px; color: rgb(119, 121, 124); font-weight: 700;' and text()='Discipline']"
    disc_display_list = "//div[@class='v-list v-select-list v-sheet theme--light v-list--dense theme--light' and @role='listbox']"

    def __init__(self, driver):
        self.driver = driver

    # logger = LogGen.loggen()

    # Selecting and validating the selected filters
    def selecting_sent_filters(self):
        try:
            self.driver.switch_to.default_content()
            frame_mail = WebDriverWait(self.driver, 10).until(expected_conditions.presence_of_element_located(
                (By.XPATH, self.mail_management_iframe_xpath)))
            self.driver.switch_to.frame(frame_mail)
            locate_filter = WebDriverWait(self.driver, 50).until(expected_conditions.presence_of_element_located(
                (By.XPATH, self.filter_button)))
            locate_filter.click()
            time.sleep(1)

            out_filter = WebDriverWait(self.driver, 10).until(expected_conditions.presence_of_element_located(
                (By.XPATH, self.filter_value_outstanding)))

            # # self.logger.info(" html info is %s", out_filter)

            # Check if the filter element is selected (active)
            out_filter_select = out_filter.get_attribute("attr-selected")
            if out_filter_select != 'true':
                print("Filter is not active for OUT")
                out_filter.click()
            else:
                print("Filter is active for OUT")
            time.sleep(1)

            na_filter = WebDriverWait(self.driver, 10).until(
                expected_conditions.presence_of_element_located((By.XPATH, self.filter_value_NA)))
            # Check if the filter element is selected (active)
            na_filter_select = na_filter.get_attribute("attr-selected")
            if na_filter_select != 'true':
                print("Filter is not active for NA")
                na_filter.click()
            else:
                print("Filter is active for NA")
            time.sleep(1)

            sent_filter = WebDriverWait(self.driver, 10).until(expected_conditions.presence_of_element_located(
                (By.XPATH, self.filter_value_Sent)))

            # Check if the filter element is selected (active)
            sent_filter_select = sent_filter.get_attribute("attr-selected")
            if sent_filter_select != 'true':
                print("Filter is not active for SENT")
                sent_filter.click()
            else:
                print("Filter is active for SENT")
            time.sleep(2)

            WebDriverWait(self.driver, 10).until(expected_conditions.presence_of_element_located(
                (By.XPATH, self.close_filter_window_event))).click()
            time.sleep(2)
        except Exception as e:
            raise e

    def validate_tra_and_click(self, traSelect):
        try:
            tra_doc_xpath = "//tr[@type='IMP_TransmittalTask']//div[@type='IMP_TransmittalTask']/div[@class='td-first-div wrap' and text()='" + traSelect + "']"

            # Scroll the element into view
            doc_element = self.driver.find_element(By.XPATH, tra_doc_xpath)
            ActionChains(self.driver).move_to_element(doc_element).perform()

            WebDriverWait(self.driver, 10).until(EC.element_to_be_clickable((By.XPATH, tra_doc_xpath))).click()

            # Right-click on the element
            ActionChains(self.driver).context_click(doc_element).perform()
            time.sleep(2)

            open_button = WebDriverWait(self.driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, self.tra_open_button)))
            open_button.click()

        except Exception as e:
            raise e

    def scroll_tra_properties(self):
        try:
            scroll_to_status = "//div[@class='scetion']//div[@title='Status' and text()='Status :']"
            scroll_to_attachments = "//button[contains(text(), 'Document Attachments')]"

            WebDriverWait(self.driver, 10).until(EC.presence_of_element_located((By.XPATH, self.mail_properties_label)))
            time.sleep(2)
            # Scroll the element into view
            doc_element = self.driver.find_element(By.XPATH, scroll_to_status)
            ActionChains(self.driver).move_to_element(doc_element).perform()

            time.sleep(3)

            doc_element = self.driver.find_element(By.XPATH, scroll_to_attachments)
            ActionChains(self.driver).move_to_element(doc_element).perform()

            time.sleep(3)

        except Exception as e:
            raise e

    def click_reply_command(self):

        try:
            reply_command_xpath = "//div[@class='floatingPanel_Header']//span[@class='floatingPanel_title']//span[@class='fa fa-reply close']"

            reply_button = WebDriverWait(self.driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, reply_command_xpath)))
            reply_button.click()

        except Exception as e:
            raise e

    def tra_Reply_ccUser(self, value):
        try:
            # uncomment below only when using TO user

            # self.driver.switch_to.default_content()
            # frame1 = WebDriverWait(self.driver, 10).until(expected_conditions.presence_of_element_located(
            #     (By.XPATH, self.mail_management_iframe_xpath)))
            # self.driver.switch_to.frame(frame1)

            WebDriverWait(self.driver, 10).until(expected_conditions.element_to_be_clickable(
                (By.XPATH, self.tra_cc_list))).click()
            time.sleep(15)
            self.driver.switch_to.default_content()
            WebDriverWait(self.driver, 20).until(expected_conditions.element_to_be_clickable(
                (By.XPATH,
                 "//div[@class='wux-controls-abstract wux-controls-responsivetileview']//div[@class='wux-controls-responsivetileview-maincontent']//div[@class='wux-button-label-ellipsis-wrap wux-ui-is-rendered wux-controls-responsivetileview-description']//span[@class='searchItemSpan' and text()='" + value + "']"))).click()
            # self.logger.info("**** Template selected ")
            WebDriverWait(self.driver, 5).until(expected_conditions.element_to_be_clickable(
                (By.XPATH, self.tra_person_ok_button))).click()
            time.sleep(5)
        except Exception as e:
            raise e

    def tra_reply(self):

        # Method to click on reply button
        try:
            self.driver.switch_to.default_content()
            frameMail = WebDriverWait(self.driver, 10).until(expected_conditions.presence_of_element_located(
                (By.XPATH, self.mail_management_iframe_xpath)))
            self.driver.switch_to.frame(frameMail)

            reply_button = WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.XPATH, self.reply_button_xpath)))

            reply_button.click()

            # success_msg = WebDriverWait(self.driver, 10).until(
            #     EC.presence_of_element_located((By.XPATH, self.tra_success_msg)))
            # # # self.logger.info(" success msg located", success_msg)
            # time.sleep(3)
            #
            # self.dp.check_doc_properties(docTitle)

            time.sleep(10)
        except Exception as e:
            raise e

    def tra_reply_ResponseRequired(self, value):
        try:

            self.driver.switch_to.default_content()
            frame_mail = WebDriverWait(self.driver, 10).until(expected_conditions.presence_of_element_located(
                (By.XPATH, self.mail_management_iframe_xpath)))
            self.driver.switch_to.frame(frame_mail)

            WebDriverWait(self.driver, 5).until(expected_conditions.element_to_be_clickable(
                (By.XPATH, "//select[@name='IMP_MailResponseRequired']//option[@value='" + value + "']"))).click()
        except Exception as e:
            raise e

    def tra_reply_Message(self, value):

        try:
            self.driver.switch_to.default_content()
            frame1 = WebDriverWait(self.driver, 10).until(expected_conditions.presence_of_element_located(
                (By.XPATH, self.mail_management_iframe_xpath)))
            self.driver.switch_to.frame(frame1)
            # self.logger.info("in frame 1")

            message_field = WebDriverWait(self.driver, 20).until(
                EC.presence_of_element_located((By.XPATH, self.locate_mesg_field)))

            self.driver.execute_script("arguments[0].scrollIntoView(true);", message_field)

            frame2 = WebDriverWait(self.driver, 10).until(expected_conditions.presence_of_element_located(
                (By.XPATH, "//div[@class='cke_contents cke_reset']//iframe[@class='cke_wysiwyg_frame cke_reset']")))
            self.driver.switch_to.frame(frame2)

            # self.logger.info("in frame 2")
            message = WebDriverWait(self.driver, 10).until(expected_conditions.presence_of_element_located(
                (By.XPATH,
                 "//html[@dir='ltr']//body[@class='cke_editable cke_editable_themed cke_contents_ltr cke_show_borders']")))
            # self.logger.info("in Message field")
            # message.send_keys("Auto Replied")
            message.send_keys(value)
            time.sleep(10)
        except Exception as e:
            raise e

    # Verifying the TRANSMITAL & Replying to a transmittal - END ===========================================================

    # Verifying the TRANSMITAL & forwarding to a transmittal - START ===========================================================

    def click_forward_command(self):

        try:
            fwd_command_xpath = "//span[@class='fa fa-mail-forward close' and @title='Forward']"

            fwd_button = WebDriverWait(self.driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, fwd_command_xpath)))
            fwd_button.click()

        except Exception as e:
            raise e

    def tra_fwd_toUser(self, username):
        try:

            WebDriverWait(self.driver, 10).until(expected_conditions.element_to_be_clickable(
                (By.XPATH, self.tra_to_list))).click()
            time.sleep(15)
            self.driver.switch_to.default_content()
            WebDriverWait(self.driver, 20).until(expected_conditions.element_to_be_clickable(
                (By.XPATH,
                 "//div[@class='wux-controls-abstract wux-controls-responsivetileview']//div[@class='wux-controls-responsivetileview-maincontent']//div[@class='wux-button-label-ellipsis-wrap wux-ui-is-rendered wux-controls-responsivetileview-description']//span[@class='searchItemSpan' and text()='" + username + "']"))).click()
            # self.logger.info("**** forward Person selected ")
            WebDriverWait(self.driver, 5).until(expected_conditions.element_to_be_clickable(
                (By.XPATH, self.tra_person_ok_button))).click()
            time.sleep(5)
        except Exception as e:
            raise e

    def tra_forward(self):

        # Method to click on reply button
        try:
            self.driver.switch_to.default_content()
            frameMail = WebDriverWait(self.driver, 10).until(expected_conditions.presence_of_element_located(
                (By.XPATH, self.mail_management_iframe_xpath)))
            self.driver.switch_to.frame(frameMail)

            fwd_button = WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.XPATH, self.forward_button_xpath)))

            fwd_button.click()

            time.sleep(10)
        except Exception as e:
            raise e

    def tra_fwd_discipline(self, discipline_value):
        try:
            desc_field = WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.XPATH, self.locate_discipline_field)))
            self.driver.execute_script("arguments[0].scrollIntoView(true);", desc_field)

            discipline_select = WebDriverWait(self.driver, 10).until(
                expected_conditions.element_to_be_clickable((By.XPATH, self.tra_discipline_field_click)))
            discipline_select.click()

            # Locate the list of disciplines
            WebDriverWait(self.driver, 30).until(expected_conditions.visibility_of_element_located(
                (By.XPATH, self.disc_display_list)))

            # self.logger.info("discipline value is %s", discipline_value)
            discipline = str(discipline_value).strip()  # Ensure discipline_value is converted to a string
            # self.logger.info("discipline value is %s", discipline)
            discipline_link_xpath = "//div[@class='v-list-item__title' and text()='" + discipline + "']"
            discipline_link_element = WebDriverWait(self.driver, 10).until(
                EC.visibility_of_element_located((By.XPATH, discipline_link_xpath)))

            # Scroll the element into view
            ActionChains(self.driver).move_to_element(discipline_link_element).perform()

            # Click the checkbox
            checkbox_xpath = "//div[@class='v-list-item__title' and text()='" + discipline + "']/../..//div[@class='v-simple-checkbox']"
            WebDriverWait(self.driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, checkbox_xpath))).click()

            time.sleep(1)
            # Close the discipline selection
            disc_click_event_xpath = "//div[contains(text(), 'Discipline') and @style='font-size: 14px; top: 3px; position: relative; padding: 4px; text-overflow: ellipsis; overflow: hidden; height: 30px; color: rgb(119, 121, 124); font-weight: 700;']"
            discipline_close_selection = WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.XPATH, disc_click_event_xpath)))

            # actions = ActionChains(self.driver)
            # actions.click(discipline_close_selection).perform()

            discipline_close_selection.click()

        except Exception as e:
            raise e

    def selecting_Inbox_filters(self):
        try:
            self.driver.switch_to.default_content()
            frame_mail = WebDriverWait(self.driver, 10).until(expected_conditions.presence_of_element_located(
                (By.XPATH, self.mail_management_iframe_xpath)))
            self.driver.switch_to.frame(frame_mail)
            locate_filter = WebDriverWait(self.driver, 50).until(expected_conditions.presence_of_element_located(
                (By.XPATH, self.filter_button)))
            locate_filter.click()
            time.sleep(1)

            out_filter = WebDriverWait(self.driver, 10).until(expected_conditions.presence_of_element_located(
                (By.XPATH, self.filter_value_outstanding)))

            # # self.logger.info(" html info is %s", out_filter)

            # Check if the filter element is selected (active)
            out_filter_select = out_filter.get_attribute("attr-selected")
            if out_filter_select != 'true':
                print("Filter is not active for OUT")
                out_filter.click()
            else:
                print("Filter is active for OUT")
            time.sleep(1)

            na_filter = WebDriverWait(self.driver, 10).until(
                expected_conditions.presence_of_element_located((By.XPATH, self.filter_value_NA)))
            # Check if the filter element is selected (active)
            na_filter_select = na_filter.get_attribute("attr-selected")
            if na_filter_select != 'true':
                print("Filter is not active for NA")
                na_filter.click()
            else:
                print("Filter is active for NA")
            time.sleep(1)

            inbox_filter = WebDriverWait(self.driver, 10).until(expected_conditions.presence_of_element_located(
                (By.XPATH, self.filter_value_Inbox)))

            # Check if the filter element is selected (active)
            inbox_filter_select = inbox_filter.get_attribute("attr-selected")
            if inbox_filter_select != 'true':
                print("Filter is not active for INBOX")
                inbox_filter.click()
            else:
                print("Filter is active for INBOX")
            time.sleep(2)

            WebDriverWait(self.driver, 10).until(expected_conditions.presence_of_element_located(
                (By.XPATH, self.close_filter_window_event))).click()
            time.sleep(2)
        except Exception as e:
            raise e

    # Verifying the TRANSMITAL & forwarding to a transmittal - END ===========================================================

    # Create a Placeholder -------------------------------------------- START


class placeholder_creation:
    plus_menu_button = "//i[@class='v-icon notranslate mdi mdi-plus theme--dark']"
    create_pH_button_xpath = "//div[@style='transition-delay: 0.05s;']//div[@class='v-responsive__content']"
    Doc_Type_xpath = "//div[@class='v-select__slot']//label[text()='Document Type *']"
    ph_serial_xpath = "//div[@class='v-text-field__slot']//label[text()='Serial Number *']/following-sibling::input[@type='text']"
    ph_title_xpath = "//div[@class='v-text-field__slot']//label[text()='Title *']/following-sibling::input[@type='text']"
    ph_locate_region_xpath = "//div[@class='v-select__slot']//label[text()='Region *']"
    ph_locate_organisation_xpath = "//div[@class='v-select__slot']//label[text()='Organization *']"
    ph_locate_discipline_xpath = "//div[@class='v-select__slot']//label[text()='Discipline *']"
    ph_discpline_list = "//div[@class='v-menu__content theme--light v-menu__content--fixed menuable__content__active']//div[@role='listbox']"
    ph_wo_list = "//div[@class='v-menu__content theme--light v-menu__content--fixed menuable__content__active']//div[@role='listbox']"
    ph_stage_list = "//div[@class='v-menu__content theme--light v-menu__content--fixed menuable__content__active']//div[@role='listbox']"
    ph_locate_wo_field = "//div[@class='v-select__slot']//label[text()='Work Order/Service Order *']"
    ph_locate_stage_field = "//div[@class='v-select__slot']//label[text()='Stage *']"
    ph_locate_bookmark_button = "//button[@class='v-btn v-btn--outlined theme--light v-size--default indigo--text']//span[text()=' Select Bookmark * ']"
    ph_locate_howmany_field = "//div[@class='v-text-field__slot']//label[text()='How Many *']/following-sibling::input[@type='number']"
    create_button_xpath = "//button[@class='ma-4 v-btn v-btn--is-elevated v-btn--has-bg theme--light v-size--default primary' and @type='button']/span[text()='Create']"
    rev_panel_close = "//button[@class='v-btn v-btn--text theme--light v-size--default secondary--text']//span[@class='v-btn__content' and text()='Close']"
    click_reviseph_button = "//ul[@class='custom-menu']//li[@title='Revise Placeholder']//div[text()='Revise Placeholder']"
    iframeNMDocumentCntrl_xpath = "//div[@id='7zAw5w30gCwH0QgEjm01']//iframe[@id='frame-9sMywGd0g06qZr2-U046']"
    select_files_xpath = "//div[@class='v-text-field__slot']//label[contains(text(),'Select File')]//following-sibling::input[@single]"
    # complete_upload = "//div[@role='progressbar']//i[contains(.,'100%')]"
    ph_name_revise_docname = "//div[@class='v-card__subtitle']//span[text()]"
    ph_locate_status_field = "//div[@class='v-select__slot']//label[text()='Status *']"
    ph_status_list = "//div[@class='v-menu__content theme--light v-menu__content--fixed menuable__content__active']//div[@role='listbox']"
    revise_button_xpath = "//button[@class='ma-4 v-btn v-btn--is-elevated v-btn--has-bg theme--light v-size--default primary']//span[text()='Revise']"
    click_cancelph_button = "//ul[@class='custom-menu']//li[@title='Cancel Placeholder']//div[text()='Cancel Placeholder']"

    def __init__(self, driver):
        self.driver = driver

    # logger = LogGen.loggen()

    def click_PH_button(self):
        try:
            WebDriverWait(self.driver, 10).until(expected_conditions.element_to_be_clickable(
                (By.XPATH, self.plus_menu_button))).click()
            WebDriverWait(self.driver, 10).until(expected_conditions.element_to_be_clickable(
                (By.XPATH, self.create_pH_button_xpath))).click()
        except Exception as e:
            raise e

    def ph_DocType(self, docType, number):
        try:
            type_selection = WebDriverWait(self.driver, 20).until(EC.presence_of_element_located(
                (By.XPATH, "//div[@class='v-select__slot']//label[text()='Document Type *']")))

            # Click using JavaScript to avoid interception
            self.driver.execute_script("arguments[0].click();", type_selection)

            # Wait for the list of document types to be visible
            WebDriverWait(self.driver, 20).until(EC.visibility_of_element_located(
                (By.XPATH,
                 "//div[@class='v-menu__content theme--light v-menu__content--fixed menuable__content__active']//div[@role='listbox']")))

            # Construct the XPath for the specific document type
            doctype_xpath = "//div[@class='v-list-item__title' and text()='" + docType + "']"

            # Wait for the element to be visible
            docType_element = WebDriverWait(self.driver, 20).until(EC.visibility_of_element_located(
                (By.XPATH, doctype_xpath)))

            # Scroll the element into view using JavaScript (alternative approach)
            # self.driver.execute_script("arguments[0].scrollIntoViewIfNeeded(true);", docType_element)
            # Scroll the element into view using JavaScript (alternative approach)
            self.driver.execute_script(
                "arguments[0].scrollIntoView({ behavior: 'smooth', block: 'center', inline: 'center' });",
                docType_element)

            # Click on the document type element
            docType_element.click()

            time.sleep(2)

            # Check if docType starts with "MDL" or "DWG" to decide whether to run the serial number code
            if docType == 'MDL - BIM Model' or docType == 'DRG - Drawing':
                # self.logger.info("inside the docType code")
                self.ph_serial(number)  # Run the serial number code

        except Exception as e:
            raise e

    def ph_serial(self, number):
        # Method to input serial details in PH creation window
        try:
            # Wait for the element to be clickable and visible
            serial_input = WebDriverWait(self.driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, self.ph_serial_xpath))
            )

            # Scroll to the element
            self.driver.execute_script("arguments[0].scrollIntoView();", serial_input)

            # Ensure the element is visible and enabled
            if serial_input.is_displayed() and serial_input.is_enabled():
                # Clear any existing text and then send keys
                serial_input.clear()
                serial_input.send_keys(number)
            else:
                raise Exception("Serial input is not visible or enabled.")

        except Exception as e:
            raise e

    def ph_title(self, title):
        # Method to input Title details in PH creation window
        try:
            # Wait for the element to be clickable
            title_input = WebDriverWait(self.driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, self.ph_title_xpath))
            )

            # Scroll to the element
            self.driver.execute_script("arguments[0].scrollIntoView();", title_input)

            # Clear any existing text and then send keys
            title_input.click()
            title_input.send_keys(title)

        except Exception as e:
            raise e

    def ph_region(self, region):
        try:
            region_field = WebDriverWait(self.driver, 10).until(EC.presence_of_element_located(
                (By.XPATH, self.ph_locate_region_xpath)))

            # Click using JavaScript to avoid interception
            self.driver.execute_script("arguments[0].click();", region_field)

            # Wait for the list of document types to be visible
            WebDriverWait(self.driver, 10).until(EC.visibility_of_element_located(
                (By.XPATH,
                 "//div[@class='v-menu__content theme--light v-menu__content--fixed menuable__content__active']//div[@role='listbox']")))

            # Construct the XPath for the specific document type
            region_xpath = "//div[@class='v-list-item v-list-item--link theme--light']//div[@class='v-list-item__content']//div[text()='" + region + "']"
            region_element = WebDriverWait(self.driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, region_xpath)))

            # Scroll the element into view
            ActionChains(self.driver).move_to_element(region_element).perform()

            # Click on the document type element
            region_element.click()

        except Exception as e:
            raise e

    def ph_organisation(self, organisation):

        try:
            org_field = WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.XPATH, self.ph_locate_organisation_xpath)))

            self.driver.execute_script("arguments[0].click();", org_field)

            # Wait for the list of document types to be visible
            WebDriverWait(self.driver, 10).until(EC.visibility_of_element_located(
                (By.XPATH,
                 "//div[@class='v-menu__content theme--light v-menu__content--fixed menuable__content__active']//div[@role='listbox']")))

            org_xpath = "//div[@class='v-list-item v-list-item--link theme--light']//div[@class='v-list-item__content']//div[text()='" + organisation + "']"
            org_element = WebDriverWait(self.driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, org_xpath)))

            # Scroll the element into view
            ActionChains(self.driver).move_to_element(org_element).perform()

            # Click on the document type element
            org_element.click()

        except Exception as e:
            raise e

    def ph_discipline(self, discipline):
        try:
            desc_field = WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.XPATH, self.ph_locate_discipline_xpath)))

            self.driver.execute_script("arguments[0].click();", desc_field)

            # Locate the list of disciplines
            WebDriverWait(self.driver, 30).until(expected_conditions.visibility_of_element_located(
                (By.XPATH, self.ph_discpline_list)))

            discipline_link_xpath = "//div[@class='v-list-item v-list-item--link theme--light']//div[@class='v-list-item__content']//div[text()='" + discipline + "']"
            discipline_link_element = WebDriverWait(self.driver, 10).until(
                EC.visibility_of_element_located((By.XPATH, discipline_link_xpath)))

            # Scroll the element into view
            ActionChains(self.driver).move_to_element(discipline_link_element).perform()

            # Click on the document type element
            discipline_link_element.click()

        except Exception as e:
            raise e

    def ph_wo(self, wo):

        try:
            wo_field = WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.XPATH, self.ph_locate_wo_field)))
            self.driver.execute_script("arguments[0].click();", wo_field)

            WebDriverWait(self.driver, 30).until(expected_conditions.visibility_of_element_located(
                (By.XPATH, self.ph_wo_list)))

            wo_select = "//div[@class='v-list-item v-list-item--link theme--light']//div[@class='v-list-item__content']//div[text()='" + wo + "']"

            wo_link_element = WebDriverWait(self.driver, 10).until(
                EC.visibility_of_element_located((By.XPATH, wo_select)))

            # Scroll the element into view
            ActionChains(self.driver).move_to_element(wo_link_element).perform()

            # Click on the document type element
            wo_link_element.click()

        except Exception as e:
            raise e

    def ph_stage(self, stage):

        try:
            stage_field = WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.XPATH, self.ph_locate_stage_field)))
            self.driver.execute_script("arguments[0].click();", stage_field)

            WebDriverWait(self.driver, 30).until(expected_conditions.visibility_of_element_located(
                (By.XPATH, self.ph_stage_list)))

            stage_select = "//div[@class='v-list-item v-list-item--link theme--light']//div[@class='v-list-item__content']//div[text()='" + stage + "']"

            stage_link_element = WebDriverWait(self.driver, 10).until(
                EC.visibility_of_element_located((By.XPATH, stage_select)))

            # Scroll the element into view
            ActionChains(self.driver).move_to_element(stage_link_element).perform()

            # Click on the document type element
            stage_link_element.click()

        except Exception as e:
            raise e

    def ph_howMany(self, number):
        try:
            how_field = WebDriverWait(self.driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, self.ph_locate_howmany_field)))

            self.driver.execute_script("arguments[0].scrollIntoView();", how_field)

            # Click on the field to ensure it has focus
            how_field.click()

            # Simulate backspace event to clear the field
            how_field.send_keys(Keys.BACKSPACE * len(how_field.get_attribute('value')))

            # Send the new value
            how_field.send_keys(number)

        except Exception as e:
            raise e

    def ph_bookmark(self):

        try:
            bmk_button = WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.XPATH, self.ph_locate_bookmark_button)))
            self.driver.execute_script("arguments[0].click();", bmk_button)

            # Scroll the element into view
            ActionChains(self.driver).move_to_element(bmk_button).perform()

        except Exception as e:
            raise e

    def ph_createButton(self):

        # Method to click on reply button
        try:
            create_button = WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.XPATH, self.create_button_xpath)))

            create_button.click()

            time.sleep(10)
        except Exception as e:
            raise e

    def validate_ph_and_revision(self, phTitle):
        # Method to select a document from reg doc tab and create WF
        try:
            ph_link_xpath = "//div[@class='v-data-table__wrapper']//tr//td//span[text()='" + phTitle + "']"
            revision_xpath = "//div[@class='v-data-table__wrapper']//tr//td//span[text()='" + phTitle + "']/../..//span[@style='cursor: pointer; color: blue;']"
            doc_name_copy_path = WebDriverWait(self.driver, 5).until(expected_conditions.visibility_of_element_located(
                (By.XPATH,
                 "//div[@class='v-data-table__wrapper']//tr//td//span[text()='" + phTitle + "']/../..//td//i[@class='v-icon notranslate v-icon--dense mdi mdi-file theme--light']/following-sibling::a[text()]")))

            # Scroll the element into view
            doc_element = self.driver.find_element(By.XPATH, ph_link_xpath)
            ActionChains(self.driver).move_to_element(doc_element).perform()

            # ============ to save the document name within the input EXCEL file for validation later - START====
            ph_name_text = doc_name_copy_path.text
            file_path = "D://Git//test-automation//feature//Code_merge//TestData//DataManager.xlsx"
            # Load the existing workbook
            workbook = openpyxl.load_workbook(file_path)
            # Select the specific sheet by name
            sheet = workbook["Placeholder"]
            # Write the extracted text to cell D3
            sheet.cell(row=3, column=12, value=ph_name_text)
            # Save the changes to the workbook
            workbook.save(file_path)
            time.sleep(1)
            # ============ to save the document name within the input EXCEL file for vlaidation later - END====

            WebDriverWait(self.driver, 10).until(EC.element_to_be_clickable((By.XPATH, revision_xpath))).click()
            time.sleep(5)

            WebDriverWait(self.driver, 5).until(expected_conditions.element_to_be_clickable(
                (By.XPATH, self.rev_panel_close))).click()

        except Exception as e:
            raise e

    # Create a Placeholder -------------------------------------------- END

    # Revise a Placeholder -------------------------------------------- START

    def select_click_ph_to_revise(self, phTitle):
        # Method to select a document from reg doc tab and click on revise PH
        try:
            revise_doc_xpath = "//div[@class='v-data-table__wrapper']//tr//td//span[text()='" + phTitle + "']/../..//td//i[@class='v-icon notranslate v-icon--dense mdi mdi-file theme--light']/following-sibling::a[@href='#']"

            # Scroll the element into view
            doc_element = self.driver.find_element(By.XPATH, revise_doc_xpath)
            ActionChains(self.driver).move_to_element(doc_element).perform()

            # Right-click on the document element
            ActionChains(self.driver).context_click(doc_element).perform()
            time.sleep(2)

            revisePH = WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.XPATH, self.click_reviseph_button)))
            revisePH.click()

        except Exception as e:
            raise e

    def ph_select_file(self, org_filepath):

        # Method to select files from local
        try:
            file_path = "D://Git//test-automation//feature//Code_merge//TestData//DataManager.xlsx"

            # Load the existing workbook
            workbook = openpyxl.load_workbook(file_path)
            # Select the specific sheet by name
            sheet = workbook["Placeholder"]

            # Read the text from cell D3
            ph_name_text = sheet.cell(row=3, column=12).value

            # ph_name_text = ph_name_value.text
            ph_name_text_updated = f"{ph_name_text}_A"
            # print("updated name is ", ph_name_text_updated)

            # Get the filename from the original filepath
            local_filename = os.path.basename(org_filepath)
            # print("orig name is ", local_filename)

            # Trim org_filepath to include only the path until "Original_upload"
            trimmed_path = os.path.dirname(org_filepath)
            while not trimmed_path.endswith("Original_upload"):
                trimmed_path = os.path.dirname(trimmed_path)

            # Generate the new filepath with the copied text and original file extension
            new_file = os.path.join(trimmed_path, f"{ph_name_text_updated}.{local_filename.split('.')[-1]}")
            print("new name is ", new_file)

            # Copy the file to the new location with the renamed filename
            shutil.copy(org_filepath, new_file)

            # Upload the updated file
            WebDriverWait(self.driver, 10).until(expected_conditions.presence_of_element_located(
                (By.XPATH, self.select_files_xpath))).send_keys(new_file)

        except Exception as e:
            raise e

    def ph_Status(self, status):

        try:
            status_field = WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.XPATH, self.ph_locate_status_field)))
            self.driver.execute_script("arguments[0].click();", status_field)

            WebDriverWait(self.driver, 30).until(expected_conditions.visibility_of_element_located(
                (By.XPATH, self.ph_status_list)))

            status_select = "//div[@class='v-list-item v-list-item--link theme--light']//div[@class='v-list-item__content']//div[text()='" + status + "']"

            status_link_element = WebDriverWait(self.driver, 10).until(
                EC.visibility_of_element_located((By.XPATH, status_select)))

            # Scroll the element into view
            ActionChains(self.driver).move_to_element(status_link_element).perform()

            # Click on the document type element
            status_link_element.click()

        except Exception as e:
            raise e

    def ph_reviseButton(self):

        # Method to click on reply button
        try:
            revise_button = WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.XPATH, self.revise_button_xpath)))

            revise_button.click()

            time.sleep(10)
        except Exception as e:
            raise e

    def validate_revised_ph(self, phName):
        # Method to select a document from reg doc tab and validate if its revised
        try:
            ph_link_xpath = "//a[contains(@href, '#') and text() = '" + phName + "']"
            ph_revision_xpath = "//a[contains(@href, '#') and text() = '" + phName + "']/../..//span[@style='cursor: pointer; color: blue;']"

            # Scroll the element into view
            doc_element = self.driver.find_element(By.XPATH, ph_link_xpath)
            ActionChains(self.driver).move_to_element(doc_element).perform()

            WebDriverWait(self.driver, 10).until(EC.element_to_be_clickable((By.XPATH, ph_revision_xpath))).click()
            time.sleep(5)

            WebDriverWait(self.driver, 5).until(expected_conditions.element_to_be_clickable(
                (By.XPATH, self.rev_panel_close))).click()

        except Exception as e:
            raise e

    # Revise a Placeholder -------------------------------------------- END

    # CANCEL a Placeholder -------------------------------------------- START

    def select_click_ph_to_cancel(self, phTitle):
        # Method to select a document from reg doc tab and cancel the PH
        try:
            revise_doc_xpath = "//div[@class='v-data-table__wrapper']//tr//td//span[text()='" + phTitle + "']/../..//td//i[@class='v-icon notranslate v-icon--dense mdi mdi-file theme--light']/following-sibling::a[@href='#']"

            # Scroll the element into view
            doc_element = self.driver.find_element(By.XPATH, revise_doc_xpath)
            ActionChains(self.driver).move_to_element(doc_element).perform()

            # Right-click on the document element
            ActionChains(self.driver).context_click(doc_element).perform()
            time.sleep(2)

            cancelPH = WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.XPATH, self.click_cancelph_button)))
            cancelPH.click()
            time.sleep(2)

            cancelPH_yesbutton = WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.XPATH,
                                                "//button[@class='v-btn v-btn--text theme--light v-size--default blue--text text--darken-1']//span[text()='Yes']")))
            cancelPH_yesbutton.click()

        except Exception as e:
            raise e

    def validate_cancelled_ph(self, phName):
        # Method to select a document from reg doc tab and validate if its cancelled
        try:
            ph_link_xpath = "//a[contains(@href, '#') and text() = '" + phName + "']"
            ph_revision_xpath = "//a[contains(@href, '#') and text() = '" + phName + "']/../..//span[@style='cursor: pointer; color: blue;']"

            # Scroll the element into view
            doc_element = self.driver.find_element(By.XPATH, ph_link_xpath)
            ActionChains(self.driver).move_to_element(doc_element).perform()

            WebDriverWait(self.driver, 10).until(EC.element_to_be_clickable((By.XPATH, ph_revision_xpath))).click()
            time.sleep(5)

            # Assuming the revision details are now visible, extract the text from the revision element
            revision_element = self.driver.find_element(By.XPATH, ph_revision_xpath)
            revision_text = revision_element.text

            # Check if the revision contains "XX"
            if "XX" in revision_text:
                print("Validation Passed: The revision contains 'XX'.")
            else:
                print("Validation Failed: The revision does not contain 'XX'.")

            time.sleep(1)

            WebDriverWait(self.driver, 5).until(expected_conditions.element_to_be_clickable(
                (By.XPATH, self.rev_panel_close))).click()

        except Exception as e:
            raise e

    # CANCEL a Placeholder -------------------------------------------- END


class iframes:  # all iframes locators

    iframeNMDocumentCntrl_xpath = "//div[@id='7zAw5w30gCwH0QgEjm01']//iframe[@id='frame-9sMywGd0g06qZr2-U046']"
    mail_management_iframe_xpath = "//div[@class='moduleWrapper']//iframe[@id='frame-9wOjNo70g06qNYY8Lm01']"
    tabDocumentUpload_xpath = "//a[contains(.,'Document Upload')]"
    tabTemporaryDocument_xpath = "//a[contains(.,'Temporary Document')]"
    tabRegisteredDocument_xpath = "//a[contains(.,'Registered Document')]"
    tabLegacyDocument_xpath = "//a[contains(.,'Legacy Document')]"
    select_files_xpath = "//div[@id='mass_upload']//div[contains(.,'Select File/s *')]//child::input[@id='input-19']"
    complete_upload = "//div[@role='progressbar']//i[contains(.,'100%')]"
    bt_inline_edit_du = "//button[@class='v-icon notranslate v-icon--link mdi mdi-pencil theme--light']"
    bt_mass_edit_du = "//div[@class='v-speed-dial v-speed-dial--bottom v-speed-dial--left v-speed-dial--direction-top']"
    bt_mass_edit_1 = "//button[@class='v-btn v-btn--is-elevated v-btn--fab v-btn--has-bg v-btn--round theme--dark v-size--small green']"
    mass_edit_form = "//div[@class='v-card__title']//span[contains(@class, 'headline') and text() = 'Mass Edit']"
    text_ms_title_field = "//label[text() = 'Title']/.."
    drp_ms_wo_field = "//label[text() = 'Work Order/Service Order']/.."
    drp_ms_status_field = "//label[text() = 'Status']/.."
    # ms_status_selection = "//div[contains(@class, 'v-list-item__title') and text() = 'For Review']"
    drp_ms_stage_field = "//label[text() = 'Stage']/.."
    # ms_stage_selection = "//div[contains(@class, 'v-list-item__title') and contains(text(), '03A')]"
    bt_update_all = "//span[contains(@class, 'v-btn__content') and text() = 'UPDATE ALL']"
    # select_a_document = "(//div[@class='v-input--selection-controls__ripple'])[3]"
    # select_a_document = "//a[contains(@href, '#') and text() = '02-ABI-BMN-3DM-000012']/../..//i[@class='v-icon notranslate mdi mdi-checkbox-blank-outline theme--light']"
    plus_menu_button = "//i[@class='v-icon notranslate mdi mdi-plus theme--dark']"
    # create_workflow_button = "//button[@class='v-btn v-btn--is-elevated v-btn--fab v-btn--has-bg v-btn--round theme--dark v-size--small yellow']//div[contains(@style, 'Create%20Workflow')]"
    widget_name = 'Document Register'
    reg_rows_per_page_filter = "//div[@class='v-input__slot' and @aria-owns='list-119']//div[@class='v-select__slot']//div[@class='v-select__selection v-select__selection--comma' and text()='50']"
    # reg_rows_per_page_filter = "//div[@class='v-data-footer__icons-after']//button[@class='v-btn v-btn--icon v-btn--round v-btn--text theme--light v-size--default']"
    reg_rows_per_page_all = "//div[@class='v-list-item__content']//div[@class='v-list-item__title' and text()='All']"
    all_rows_selected = "//div[@class='v-input__slot' and @aria-owns='list-119']//div[@class='v-select__slot']//div[@class='v-select__selection v-select__selection--comma' and text()='All']"

    create_workflow_button = "//div[@style='transition-delay: 0.2s;']//div[@class='v-responsive__content']"
    select_reasonForIssue = "//select[@name='IMP_WFReasonForIssue']/option[1]"
    wf_title = "//input[@name='IMP_WFTitle']"
    wf_create_button = "//button[@class='btn-primary btn hd-form-submit-hover hd-form-submit-focus']"
    wf_template_selector_button = "//div[@class='fa fa-search search']"
    template_page_locator = "//span[@class=' set-detail-view-title' and text()='Results']"
    template_select_from_list = "//div[@class='wux-controls-responsivetileview-maincontent']//span[@class='searchItemSpan search_item_in_apps' and text()='WK-Approve-Task-Template']"
    template_ok_button = "//div[@class='buttonContainer footer_button']//button[@id='id_in_app_ok' and text()='OK']"
    route_selection = "//div[@class='wux-tweakers wux-tweakers-urlobject']//a[@class='wux-tweakers-string-label' and text()='Not Started']"
    right_click_start_button = "//div[@class='wux-menu-cell wux-menu-push']//div[@data-wux-menu-item-uid='1']"
    iframe_Route_Management = "//div[@id='m_preview-7362fc']//iframe[@id='frame-preview-7362fc']"

    # constructor to initialise driver
    def __init__(self, driver):
        self.lp = LoginPage(driver)
        self.driver = driver
        self.dp = regDocument_properties(self.driver)

    # logger = LogGen.loggen()

    def navigate_to_tab_document_upload(self):

        # Method to navigate to the Document tab in Document Register widget

        try:
            # to get to the default iFrame
            self.driver.switch_to.default_content()
            frame1 = WebDriverWait(self.driver, 20).until(expected_conditions.presence_of_element_located(
                (By.XPATH, self.iframeNMDocumentCntrl_xpath)))
            self.driver.switch_to.frame(frame1)
            WebDriverWait(self.driver, 20).until(expected_conditions.element_to_be_clickable(
                (By.XPATH, self.tabDocumentUpload_xpath))).click()

        except Exception as e:
            raise e

    def navigate_to_tab_registered_document(self):

        #  Method to navigate to the Registered Document tab in Document Register widget

        try:
            self.driver.switch_to.default_content()
            frame1 = WebDriverWait(self.driver, 100).until(expected_conditions.presence_of_element_located(
                (By.XPATH, self.iframeNMDocumentCntrl_xpath)))
            self.driver.switch_to.frame(frame1)
            WebDriverWait(self.driver, 500).until(expected_conditions.presence_of_element_located(
                (By.XPATH, self.tabRegisteredDocument_xpath)))
            WebDriverWait(self.driver, 250).until(expected_conditions.element_to_be_clickable(
                (By.XPATH, self.tabRegisteredDocument_xpath))).click()
            time.sleep(15)

        except Exception as e:
            raise e

    def select_a_document(self, docTitle):
        # Method to select a document from reg doc tab and create WF
        try:
            doc_link_xpath = "//a[contains(@href, '#') and text() = '" + docTitle + "']"
            checkbox_xpath = "//a[contains(@href, '#') and text() = '" + docTitle + "']/../..//div[@class='v-input--selection-controls__ripple']"

            # WebDriverWait(self.driver, 10).until(expected_conditions.presence_of_element_located(
            #     (By.XPATH, self.all_rows_selected)))

            # Scroll the element into view
            doc_element = self.driver.find_element(By.XPATH, doc_link_xpath)
            ActionChains(self.driver).move_to_element(doc_element).perform()

            WebDriverWait(self.driver, 10).until(EC.element_to_be_clickable((By.XPATH, checkbox_xpath))).click()

        except Exception as e:
            raise e

    def create_workflow(self):
        try:

            WebDriverWait(self.driver, 10).until(expected_conditions.element_to_be_clickable(
                (By.XPATH, self.plus_menu_button))).click()
            WebDriverWait(self.driver, 10).until(expected_conditions.element_to_be_clickable(
                (By.XPATH, self.create_workflow_button))).click()
        except Exception as e:
            raise e

    def wf_creation_window(self, title):

        #  Method to input WF details in WF creation window
        try:
            WebDriverWait(self.driver, 5).until(expected_conditions.element_to_be_clickable(
                (By.XPATH, self.select_reasonForIssue))).click()
            WebDriverWait(self.driver, 5).until(expected_conditions.element_to_be_clickable(
                (By.XPATH, self.wf_title))).send_keys(title)
            try:
                WebDriverWait(self.driver, 5).until(expected_conditions.element_to_be_clickable(
                    (By.XPATH, self.wf_template_selector_button))).click()
                time.sleep(15)
                # below "default_Content" code because to search for template, need to switch the frame
                # WebDriverWait(self.driver, 5).until(expected_conditions.presence_of_element_located(
                #     (By.XPATH, self.template_page_locator)))

                self.driver.switch_to.default_content()
                WebDriverWait(self.driver, 10).until(expected_conditions.element_to_be_clickable(
                    (By.XPATH, self.template_select_from_list))).click()

                WebDriverWait(self.driver, 5).until(expected_conditions.element_to_be_clickable(
                    (By.XPATH, self.template_ok_button))).click()
                time.sleep(5)

            except Exception as e:
                raise e

            self.driver.switch_to.default_content()
            time.sleep(5)
            frame1 = WebDriverWait(self.driver, 10).until(expected_conditions.presence_of_element_located(
                (By.XPATH, self.iframeNMDocumentCntrl_xpath)))
            self.driver.switch_to.frame(frame1)
            WebDriverWait(self.driver, 5).until(expected_conditions.element_to_be_clickable(
                (By.XPATH, self.wf_create_button))).click()
            time.sleep(5)

        except Exception as e:
            raise e

    def start_workflow(self):
        try:
            time.sleep(10)
            frame1 = WebDriverWait(self.driver, 10).until(expected_conditions.presence_of_element_located(
                (By.XPATH, self.iframe_Route_Management)))
            self.driver.switch_to.default_content(frame1)
            # self.logger.info"*** frame swicthed ***")
            route_element = WebDriverWait(self.driver, 10).until(expected_conditions.element_to_be_clickable(
                (By.XPATH, self.route_selection))).click()
            ActionChains(self.driver).double_click(route_element).perform()

            # WebDriverWait(self.driver, 10).until(expected_conditions.element_to_be_clickable(
            #     (By.XPATH, self.right_click_start_button))).click()

        except Exception as e:
            raise e

    # def select_file(self, filepath):
    #     '''
    #             Method to select files from local
    #             '''
    #     try:
    #         self.driver.switch_to.default_content()
    #         frame1 = WebDriverWait(self.driver, 10).until(expected_conditions.presence_of_element_located(
    #             (By.XPATH, self.iframeNMDocumentCntrl_xpath)))
    #         self.driver.switch_to.frame(frame1)
    #         WebDriverWait(self.driver, 10).until(expected_conditions.presence_of_element_located(
    #             (By.XPATH, self.select_files_xpath))).send_keys(filepath)
    #         WebDriverWait(self.driver, 60).until(expected_conditions.presence_of_element_located(
    #             (By.XPATH, self.complete_upload)))
    #
    #     except Exception as e:
    #         raise e
    #
    # # def update_document(self, title, wo, status, stage):
    # def update_document(self, title):
    #     '''
    #             Method to wait till completion of file upload
    #             '''
    #     try:
    #         WebDriverWait(self.driver, 60).until(expected_conditions.presence_of_element_located(
    #             (By.XPATH, self.complete_upload)))
    #         WebDriverWait(self.driver, 10).until(expected_conditions.presence_of_element_located(
    #             (By.XPATH, self.bt_mass_edit_du))).click()
    #         WebDriverWait(self.driver, 10).until(expected_conditions.presence_of_element_located(
    #             (By.XPATH, self.bt_mass_edit_1))).click()
    #         WebDriverWait(self.driver, 10).until(expected_conditions.presence_of_element_located(
    #             (By.XPATH, self.mass_edit_form)))
    #         try:
    #             ''' Entering text into title field '''
    #             title_field = WebDriverWait(self.driver, 10).until(expected_conditions.presence_of_element_located(
    #                 (By.XPATH, self.text_ms_title_field)))
    #             title_field.click()
    #             time.sleep(2)
    #             title_field.send_keys(title)
    #
    #         except Exception as e:
    #             raise e
    #
    #         try:
    #             ''' selecting wo from dropdown '''
    #             ms_wo_selection = WebDriverWait(self.driver, 10).until(expected_conditions.presence_of_element_located(
    #                 (By.XPATH, self.drp_ms_wo_field)))
    #             ms_wo_selection.click()
    #             WebDriverWait(self.driver, 10).until(expected_conditions.presence_of_element_located(
    #                 (By.XPATH, self.ms_wo_list))).click()
    #
    #         except Exception as e:
    #             raise e
    #
    #         # # ms_wo_selection = Select(WebDriverWait(self.driver, 10).until(expected_conditions.presence_of_element_located(
    #         # #    (By.XPATH, self.drp_ms_wo_field))))
    #         # # ms_wo_selection.select_by_visible_text(wo)
    #         #
    #         # ms_status_selection = Select(WebDriverWait(self.driver, 10).until(expected_conditions.presence_of_element_located(
    #         #    (By.XPATH, self.drp_ms_status_field))))
    #         # ms_status_selection.select_by_visible_text(status)
    #         #
    #         # ms_stage_selection = Select(WebDriverWait(self.driver, 10).until(expected_conditions.presence_of_element_located(
    #         #    (By.XPATH, self.drp_ms_stage_field))))
    #         # ms_stage_selection.select_by_visible_text(stage)
    #         #
    #         # WebDriverWait(self.driver, 10).until(expected_conditions.presence_of_element_located(
    #         #    (By.XPATH, self.bt_update_all))).click()
    #
    #     except Exception as e:
    #         raise e

    # -------- Transmittal Management START--------------------------------------------
