import time

from selenium.webdriver import Keys

from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl
import os
import shutil

# Create a Placeholder -------------------------------------------- START

class placeholder_creation:
    plus_menu_button = "//i[@class='v-icon notranslate mdi mdi-plus theme--dark']"
    create_pH_button_xpath = "//div[@style='transition-delay: 0.05s;']//div[@class='v-responsive__content']"
    Doc_Type_xpath = "//div[@class='v-select__slot']//label[text()='Document Type *']"
    ph_serial_xpath = "//div[@class='v-text-field__slot']//label[text()='Serial Number *']/following-sibling::input[@type='text']"
    ph_title_xpath = "//div[@class='v-text-field__slot']//label[text()='Title *']/following-sibling::input[@type='text']"
    ph_locate_region_xpath = "//div[@class='v-select__slot']//label[text()='Region *']"
    ph_locate_organisation_xpath = "//div[@class='v-select__slot']//label[text()='Organization *']"
    ph_locate_discipline_xpath = "//div[@class='v-select__slot']//label[text()='Discipline *']"
    ph_discpline_list = "//div[@class='v-menu__content theme--light v-menu__content--fixed menuable__content__active']//div[@role='listbox']"
    ph_wo_list = "//div[@class='v-menu__content theme--light v-menu__content--fixed menuable__content__active']//div[@role='listbox']"
    ph_stage_list = "//div[@class='v-menu__content theme--light v-menu__content--fixed menuable__content__active']//div[@role='listbox']"
    ph_locate_wo_field = "//div[@class='v-select__slot']//label[text()='Work Order/Service Order *']"
    ph_locate_stage_field = "//div[@class='v-select__slot']//label[text()='Stage *']"
    ph_locate_bookmark_button = "//button[@class='v-btn v-btn--outlined theme--light v-size--default indigo--text']//span[text()=' Select Bookmark * ']"
    ph_locate_howmany_field = "//div[@class='v-text-field__slot']//label[text()='How Many *']/following-sibling::input[@type='number']"
    create_button_xpath = "//button[@class='ma-4 v-btn v-btn--is-elevated v-btn--has-bg theme--light v-size--default primary' and @type='button']/span[text()='Create']"
    rev_panel_close = "//button[@class='v-btn v-btn--text theme--light v-size--default secondary--text']//span[@class='v-btn__content' and text()='Close']"
    click_reviseph_button = "//ul[@class='custom-menu']//li[@title='Revise Placeholder']//div[text()='Revise Placeholder']"
    iframeNMDocumentCntrl_xpath = "//div[@id='7zAw5w30gCwH0QgEjm01']//iframe[@id='frame-9sMywGd0g06qZr2-U046']"
    select_files_xpath = "//div[@class='v-text-field__slot']//label[contains(text(),'Select File')]//following-sibling::input[@single]"
    # complete_upload = "//div[@role='progressbar']//i[contains(.,'100%')]"
    ph_name_revise_docname = "//div[@class='v-card__subtitle']//span[text()]"
    ph_locate_status_field = "//div[@class='v-select__slot']//label[text()='Status *']"
    ph_status_list = "//div[@class='v-menu__content theme--light v-menu__content--fixed menuable__content__active']//div[@role='listbox']"
    revise_button_xpath = "//button[@class='ma-4 v-btn v-btn--is-elevated v-btn--has-bg theme--light v-size--default primary']//span[text()='Revise']"
    click_cancelph_button = "//ul[@class='custom-menu']//li[@title='Cancel Placeholder']//div[text()='Cancel Placeholder']"
    ph_success_msg = "//div[@class='alert alert-success alert-dismissible' and contains(text(), 'Uploaded Successfully!')]"

    def __init__(self, driver):
        self.driver = driver

    # logger = LogGen.loggen()

    def click_PH_button(self):
        try:
            WebDriverWait(self.driver, 10).until(expected_conditions.element_to_be_clickable(
                (By.XPATH, self.plus_menu_button))).click()
            WebDriverWait(self.driver, 10).until(expected_conditions.element_to_be_clickable(
                (By.XPATH, self.create_pH_button_xpath))).click()
        except Exception as e:
            raise e

    def ph_DocType(self, docType, number):
        try:
            type_selection = WebDriverWait(self.driver, 20).until(EC.presence_of_element_located(
                (By.XPATH, "//div[@class='v-select__slot']//label[text()='Document Type *']")))

            # Click using JavaScript to avoid interception
            self.driver.execute_script("arguments[0].click();", type_selection)

            # Wait for the list of document types to be visible
            WebDriverWait(self.driver, 20).until(EC.visibility_of_element_located(
                (By.XPATH,
                 "//div[@class='v-menu__content theme--light v-menu__content--fixed menuable__content__active']//div[@role='listbox']")))

            # Construct the XPath for the specific document type
            doctype_xpath = "//div[@class='v-list-item__title' and text()='" + docType + "']"

            # Wait for the element to be visible
            docType_element = WebDriverWait(self.driver, 20).until(EC.visibility_of_element_located(
                (By.XPATH, doctype_xpath)))

            # Scroll the element into view using JavaScript (alternative approach)
            # self.driver.execute_script("arguments[0].scrollIntoViewIfNeeded(true);", docType_element)
            # Scroll the element into view using JavaScript (alternative approach)
            self.driver.execute_script(
                "arguments[0].scrollIntoView({ behavior: 'smooth', block: 'center', inline: 'center' });",
                docType_element)

            # Click on the document type element
            docType_element.click()

            time.sleep(2)

            # Check if docType starts with "MDL" or "DWG" to decide whether to run the serial number code
            if docType == 'MDL - BIM Model' or docType == 'DRG - Drawing':
                # logger_setup.logger.info("inside the docType code")
                self.ph_serial(number)  # Run the serial number code

        except Exception as e:
            raise e

    def ph_serial(self, number):
        # Method to input serial details in PH creation window
        try:
            # Wait for the element to be clickable and visible
            serial_input = WebDriverWait(self.driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, self.ph_serial_xpath))
            )

            # Scroll to the element
            self.driver.execute_script("arguments[0].scrollIntoView();", serial_input)

            # Ensure the element is visible and enabled
            if serial_input.is_displayed() and serial_input.is_enabled():
                # Clear any existing text and then send keys
                serial_input.clear()
                serial_input.send_keys(number)
            else:
                raise Exception("Serial input is not visible or enabled.")

        except Exception as e:
            raise e

    def ph_title(self, title):
        # Method to input Title details in PH creation window
        try:
            # Wait for the element to be clickable
            title_input = WebDriverWait(self.driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, self.ph_title_xpath))
            )

            # Scroll to the element
            self.driver.execute_script("arguments[0].scrollIntoView();", title_input)

            # Clear any existing text and then send keys
            title_input.click()
            title_input.send_keys(title)

        except Exception as e:
            raise e

    def ph_region(self, region):
        try:
            region_field = WebDriverWait(self.driver, 10).until(EC.presence_of_element_located(
                (By.XPATH, self.ph_locate_region_xpath)))

            # Click using JavaScript to avoid interception
            self.driver.execute_script("arguments[0].click();", region_field)

            # Wait for the list of document types to be visible
            WebDriverWait(self.driver, 10).until(EC.visibility_of_element_located(
                (By.XPATH,
                 "//div[@class='v-menu__content theme--light v-menu__content--fixed menuable__content__active']//div[@role='listbox']")))

            # Construct the XPath for the specific document type
            region_xpath = "//div[@class='v-list-item v-list-item--link theme--light']//div[@class='v-list-item__content']//div[text()='" + region + "']"
            region_element = WebDriverWait(self.driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, region_xpath)))

            # Scroll the element into view
            ActionChains(self.driver).move_to_element(region_element).perform()

            # Click on the document type element
            region_element.click()

        except Exception as e:
            raise e

    def ph_organisation(self, organisation):

        try:
            org_field = WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.XPATH, self.ph_locate_organisation_xpath)))

            self.driver.execute_script("arguments[0].click();", org_field)

            # Wait for the list of document types to be visible
            WebDriverWait(self.driver, 10).until(EC.visibility_of_element_located(
                (By.XPATH,
                 "//div[@class='v-menu__content theme--light v-menu__content--fixed menuable__content__active']//div[@role='listbox']")))

            org_xpath = "//div[@class='v-list-item v-list-item--link theme--light']//div[@class='v-list-item__content']//div[text()='" + organisation + "']"
            org_element = WebDriverWait(self.driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, org_xpath)))

            # Scroll the element into view
            ActionChains(self.driver).move_to_element(org_element).perform()

            # Click on the document type element
            org_element.click()

        except Exception as e:
            raise e

    def ph_discipline(self, discipline):
        try:
            desc_field = WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.XPATH, self.ph_locate_discipline_xpath)))

            self.driver.execute_script("arguments[0].click();", desc_field)

            # Locate the list of disciplines
            WebDriverWait(self.driver, 30).until(expected_conditions.visibility_of_element_located(
                (By.XPATH, self.ph_discpline_list)))

            discipline_link_xpath = "//div[@class='v-list-item v-list-item--link theme--light']//div[@class='v-list-item__content']//div[text()='" + discipline + "']"
            discipline_link_element = WebDriverWait(self.driver, 10).until(
                EC.visibility_of_element_located((By.XPATH, discipline_link_xpath)))

            # Scroll the element into view
            ActionChains(self.driver).move_to_element(discipline_link_element).perform()

            # Click on the document type element
            discipline_link_element.click()

        except Exception as e:
            raise e

    def ph_wo(self, wo):

        try:
            wo_field = WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.XPATH, self.ph_locate_wo_field)))
            self.driver.execute_script("arguments[0].click();", wo_field)

            WebDriverWait(self.driver, 30).until(expected_conditions.visibility_of_element_located(
                (By.XPATH, self.ph_wo_list)))

            wo_select = "//div[@class='v-list-item v-list-item--link theme--light']//div[@class='v-list-item__content']//div[text()='" + wo + "']"

            wo_link_element = WebDriverWait(self.driver, 10).until(
                EC.visibility_of_element_located((By.XPATH, wo_select)))

            # Scroll the element into view
            ActionChains(self.driver).move_to_element(wo_link_element).perform()

            # Click on the document type element
            wo_link_element.click()

        except Exception as e:
            raise e

    def ph_stage(self, stage):

        try:
            stage_field = WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.XPATH, self.ph_locate_stage_field)))
            self.driver.execute_script("arguments[0].click();", stage_field)

            WebDriverWait(self.driver, 30).until(expected_conditions.visibility_of_element_located(
                (By.XPATH, self.ph_stage_list)))

            stage_select = "//div[@class='v-list-item v-list-item--link theme--light']//div[@class='v-list-item__content']//div[text()='" + stage + "']"

            stage_link_element = WebDriverWait(self.driver, 10).until(
                EC.visibility_of_element_located((By.XPATH, stage_select)))

            # Scroll the element into view
            ActionChains(self.driver).move_to_element(stage_link_element).perform()

            # Click on the document type element
            stage_link_element.click()

        except Exception as e:
            raise e

    def ph_howMany(self, number):
        try:
            how_field = WebDriverWait(self.driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, self.ph_locate_howmany_field)))

            self.driver.execute_script("arguments[0].scrollIntoView();", how_field)

            # Click on the field to ensure it has focus
            how_field.click()

            # Simulate backspace event to clear the field
            how_field.send_keys(Keys.BACKSPACE * len(how_field.get_attribute('value')))

            # Send the new value
            how_field.send_keys(number)

        except Exception as e:
            raise e

    def ph_bookmark(self):

        try:
            bmk_button = WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.XPATH, self.ph_locate_bookmark_button)))
            self.driver.execute_script("arguments[0].click();", bmk_button)

            # Scroll the element into view
            ActionChains(self.driver).move_to_element(bmk_button).perform()

        except Exception as e:
            raise e

    def ph_createButton(self):

        # Method to click on reply button
        try:
            create_button = WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.XPATH, self.create_button_xpath)))

            create_button.click()

            success_msg = WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.XPATH, self.ph_success_msg)))
            assert success_msg.is_displayed(), "Success message not displayed after creating PH"

            time.sleep(10)
        except Exception as e:
            raise e

    def validate_ph_and_revision(self, phTitle, file_path):
        # Method to select a document from reg doc tab and create WF
        try:
            ph_link_xpath = "//div[@class='v-data-table__wrapper']//tr//td//span[text()='" + phTitle + "']"
            revision_xpath = "//div[@class='v-data-table__wrapper']//tr//td//span[text()='" + phTitle + "']/../..//span[@style='cursor: pointer; color: blue;']"
            doc_name_copy_path = WebDriverWait(self.driver, 10).until(expected_conditions.visibility_of_element_located(
                (By.XPATH,
                 "//div[@class='v-data-table__wrapper']//tr//td//span[text()='" + phTitle + "']/../..//td//i[@class='v-icon notranslate v-icon--dense mdi mdi-file theme--light']/following-sibling::a[text()]")))

            # Scroll the element into view
            doc_element = self.driver.find_element(By.XPATH, ph_link_xpath)
            ActionChains(self.driver).move_to_element(doc_element).perform()

            # ============ to save the document name within the input EXCEL file for validation later - START====
            ph_name_text = doc_name_copy_path.text
            # file_path = ".//TestData//DataManager.xlsx"
            # Load the existing workbook
            workbook = openpyxl.load_workbook(file_path)
            # Select the specific sheet by name
            sheet = workbook["Placeholder"]
            # Write the extracted text to cell D3
            sheet.cell(row=3, column=12, value=ph_name_text)
            # Save the changes to the workbook
            workbook.save(file_path)
            time.sleep(1)
            # ============ to save the document name within the input EXCEL file for vlaidation later - END====

            WebDriverWait(self.driver, 10).until(EC.element_to_be_clickable((By.XPATH, revision_xpath))).click()
            time.sleep(5)

            WebDriverWait(self.driver, 5).until(expected_conditions.element_to_be_clickable(
                (By.XPATH, self.rev_panel_close))).click()

        except Exception as e:
            raise e

    # Create a Placeholder -------------------------------------------- END

    # Revise a Placeholder -------------------------------------------- START

    def select_click_ph_to_revise(self, phTitle):
        # Method to select a document from reg doc tab and click on revise PH
        try:
            revise_doc_xpath = "//div[@class='v-data-table__wrapper']//tr//td//span[text()='" + phTitle + "']/../..//td//i[@class='v-icon notranslate v-icon--dense mdi mdi-file theme--light']/following-sibling::a[@href='#']"

            # Scroll the element into view
            doc_element = self.driver.find_element(By.XPATH, revise_doc_xpath)
            ActionChains(self.driver).move_to_element(doc_element).perform()

            # Right-click on the document element
            ActionChains(self.driver).context_click(doc_element).perform()
            time.sleep(2)

            revisePH = WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.XPATH, self.click_reviseph_button)))
            revisePH.click()

        except Exception as e:
            raise e

    def ph_select_file(self, org_filepath, file_path):

        # Method to select files from local
        try:
            # file_path = ".//TestData//DataManager.xlsx"

            # Load the existing workbook
            workbook = openpyxl.load_workbook(file_path)
            # Select the specific sheet by name
            sheet = workbook["Placeholder"]

            # Read the text from cell D3
            ph_name_text = sheet.cell(row=3, column=12).value

            # ph_name_text = ph_name_value.text
            ph_name_text_updated = f"{ph_name_text}_A"
            # print("updated name is ", ph_name_text_updated)

            # Get the filename from the original filepath
            local_filename = os.path.basename(org_filepath)
            # print("orig name is ", local_filename)

            # Trim org_filepath to include only the path until "Original_upload"
            trimmed_path = os.path.dirname(org_filepath)
            while not trimmed_path.endswith("Original_upload"):
                trimmed_path = os.path.dirname(trimmed_path)

            # Generate the new filepath with the copied text and original file extension
            new_file = os.path.join(trimmed_path, f"{ph_name_text_updated}.{local_filename.split('.')[-1]}")
            print("new name is ", new_file)

            # Copy the file to the new location with the renamed filename
            shutil.copy(org_filepath, new_file)

            # Upload the updated file
            WebDriverWait(self.driver, 10).until(expected_conditions.presence_of_element_located(
                (By.XPATH, self.select_files_xpath))).send_keys(new_file)

        except Exception as e:
            raise e

    def ph_Status(self, status):

        try:
            status_field = WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.XPATH, self.ph_locate_status_field)))
            self.driver.execute_script("arguments[0].click();", status_field)

            WebDriverWait(self.driver, 30).until(expected_conditions.visibility_of_element_located(
                (By.XPATH, self.ph_status_list)))

            status_select = "//div[@class='v-list-item v-list-item--link theme--light']//div[@class='v-list-item__content']//div[text()='" + status + "']"

            status_link_element = WebDriverWait(self.driver, 10).until(
                EC.visibility_of_element_located((By.XPATH, status_select)))

            # Scroll the element into view
            ActionChains(self.driver).move_to_element(status_link_element).perform()

            # Click on the document type element
            status_link_element.click()

        except Exception as e:
            raise e

    def ph_reviseButton(self):

        # Method to click on reply button
        try:
            revise_button = WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.XPATH, self.revise_button_xpath)))

            revise_button.click()

            success_msg = WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.XPATH, self.ph_success_msg)))
            assert success_msg.is_displayed(), "Success message not displayed after revising PH"

            time.sleep(10)
        except Exception as e:
            raise e

    def validate_revised_ph(self, phName):
        # Method to select a document from reg doc tab and validate if its revised
        try:
            ph_link_xpath = "//a[contains(@href, '#') and text() = '" + phName + "']"
            ph_revision_xpath = "//a[contains(@href, '#') and text() = '" + phName + "']/../..//span[@style='cursor: pointer; color: blue;']"

            # Scroll the element into view
            doc_element = self.driver.find_element(By.XPATH, ph_link_xpath)
            ActionChains(self.driver).move_to_element(doc_element).perform()

            WebDriverWait(self.driver, 10).until(EC.element_to_be_clickable((By.XPATH, ph_revision_xpath))).click()
            time.sleep(5)

            WebDriverWait(self.driver, 5).until(expected_conditions.element_to_be_clickable(
                (By.XPATH, self.rev_panel_close))).click()

        except Exception as e:
            raise e

    # Revise a Placeholder -------------------------------------------- END

    # CANCEL a Placeholder -------------------------------------------- START

    def select_click_ph_to_cancel(self, phTitle):
        # Method to select a document from reg doc tab and cancel the PH
        try:
            revise_doc_xpath = "//div[@class='v-data-table__wrapper']//tr//td//span[text()='" + phTitle + "']/../..//td//i[@class='v-icon notranslate v-icon--dense mdi mdi-file theme--light']/following-sibling::a[@href='#']"

            # Scroll the element into view
            doc_element = self.driver.find_element(By.XPATH, revise_doc_xpath)
            ActionChains(self.driver).move_to_element(doc_element).perform()

            # Right-click on the document element
            ActionChains(self.driver).context_click(doc_element).perform()
            time.sleep(2)

            cancelPH = WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.XPATH, self.click_cancelph_button)))
            cancelPH.click()
            time.sleep(2)

            cancelPH_yesbutton = WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.XPATH,
                                                "//button[@class='v-btn v-btn--text theme--light v-size--default blue--text text--darken-1']//span[text()='Yes']")))
            cancelPH_yesbutton.click()

        except Exception as e:
            raise e

    def validate_cancelled_ph(self, phName):
        # Method to select a document from reg doc tab and validate if its cancelled
        try:
            ph_link_xpath = "//a[contains(@href, '#') and text() = '" + phName + "']"
            ph_revision_xpath = "//a[contains(@href, '#') and text() = '" + phName + "']/../..//span[@style='cursor: pointer; color: blue;']"

            # Scroll the element into view
            doc_element = self.driver.find_element(By.XPATH, ph_link_xpath)
            ActionChains(self.driver).move_to_element(doc_element).perform()

            WebDriverWait(self.driver, 10).until(EC.element_to_be_clickable((By.XPATH, ph_revision_xpath))).click()
            time.sleep(5)

            # Assuming the revision details are now visible, extract the text from the revision element
            revision_element = self.driver.find_element(By.XPATH, ph_revision_xpath)
            revision_text = revision_element.text

            # Check if the revision contains "XX"
            if "XX" in revision_text:
                print("Validation Passed: The revision contains 'XX'.")
                assert True
            else:
                print("Validation Failed: The revision does not contain 'XX'.")
                assert False

            time.sleep(1)

            WebDriverWait(self.driver, 5).until(expected_conditions.element_to_be_clickable(
                (By.XPATH, self.rev_panel_close))).click()

        except Exception as e:
            raise e

    # CANCEL a Placeholder -------------------------------------------- END
